"""Módulo Reportes — Diseño moderno Venialgo POS"""
import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime, timedelta
from utils.formatters import format_currency
try:
    from utils.window_icon import set_icon as _set_icon
except ImportError:
    def _set_icon(w): pass


# ── Gráfico de barras con Canvas (sin dependencias externas) ─────────────────
class BarChart(tk.Canvas):
    """Gráfico de barras con esquinas redondeadas en los tops. Sin dependencias."""
    BAR_COLOR  = "#2563EB"
    GRID_COLOR = "#F3F4F6"
    LABEL_FG   = "#6B7280"

    def __init__(self, parent, **kw):
        kw.setdefault("bg", "#FFFFFF")
        kw.setdefault("highlightthickness", 0)
        kw.setdefault("height", 175)
        super().__init__(parent, **kw)
        self._data = []
        self.bind("<Configure>", lambda e: self.after_idle(self._draw))

    def set_data(self, data):
        self._data = data
        self.after_idle(self._draw)

    def _rounded_bar(self, x0, y0, x1, y1, r=5):
        r = min(r, max(1,(x1-x0)//2), max(1,(y1-y0)//2))
        c = self.BAR_COLOR
        # Body
        self.create_rectangle(x0, y0+r, x1, y1, fill=c, outline='')
        # Top flat rect (covers gap between arcs)
        self.create_rectangle(x0+r, y0, x1-r, y0+r+1, fill=c, outline='')
        # Two top arcs
        self.create_arc(x0, y0, x0+2*r, y0+2*r,
                        start=90, extent=90, fill=c, outline=c)
        self.create_arc(x1-2*r, y0, x1, y0+2*r,
                        start=0,  extent=90, fill=c, outline=c)

    def _draw(self):
        self.delete("all")
        w = self.winfo_width()
        h = self.winfo_height()
        if w < 20 or h < 20:
            return
        if not self._data:
            self.create_text(w//2, h//2, text="Sin datos para el período",
                             fill=self.LABEL_FG, font=(FONT, 9))
            return
        pad_l, pad_r, pad_t, pad_b = 10, 10, 20, 34
        n      = len(self._data)
        vals   = [v for _, v in self._data]
        max_v  = max(vals) if any(v > 0 for v in vals) else 1
        cw     = w - pad_l - pad_r
        ch     = h - pad_t - pad_b
        gap    = max(3, min(8, cw // (n * 4))) if n > 1 else 0
        bar_w  = max(6, int((cw - gap * (n - 1)) / n)) if n > 0 else cw
        # Grid lines
        for i in range(1, 4):
            y = pad_t + ch * (1 - i / 3)
            self.create_line(pad_l, y, w - pad_r, y,
                             fill=self.GRID_COLOR, dash=(4, 3))
        # Bars
        for idx, (label, val) in enumerate(self._data):
            x0 = pad_l + idx * (bar_w + gap)
            x1 = x0 + bar_w
            bh = int(ch * val / max_v) if max_v > 0 else 0
            y0 = pad_t + ch - bh
            y1 = pad_t + ch
            if bh > 0:
                if bh >= 10:
                    self._rounded_bar(x0, y0, x1, y1, r=5)
                else:
                    self.create_rectangle(x0, y0, x1, y1,
                                          fill=self.BAR_COLOR, outline='')
            if val > 0:
                try:
                    from utils.formatters import format_currency
                    lv = format_currency(val) if val >= 10000 else str(int(val))
                except Exception:
                    lv = str(int(val))
                self.create_text((x0+x1)/2, y0-4, text=lv,
                                 fill=self.LABEL_FG,
                                 font=(FONT, 7, "bold"), anchor="s")
            self.create_text((x0+x1)/2, y1+5, text=label,
                             fill=self.LABEL_FG,
                             font=(FONT, 7, "bold"), anchor="n")


THEME={"ct_bg":"#F9FAFB","card_bg":"#FFFFFF","card_border":"#E5E7EB","sb_bg":"#111827",
    "txt_primary":"#111827","txt_secondary":"#6B7280","txt_white":"#FFFFFF",
    "acc_blue":"#2563EB","acc_green":"#059669","acc_amber":"#D97706","acc_rose":"#E11D48",
    "acc_purple":"#7C3AED","acc_cyan":"#0891B2","btn_secondary":"#6B7280",
    "input_bg":"#FFFFFF","input_brd":"#D1D5DB","input_foc":"#2563EB",
    "row_odd":"#F9FAFB","row_even":"#FFFFFF"}
FONT="Segoe UI"

# ── Rounded card widget ───────────────────────────────────────────────────
def _draw_rr(c, x1, y1, x2, y2, r, fill, outline):
    """Rounded rect on canvas; tag='rc'."""
    kw = {'tags': 'rc'}
    r = min(r, max(1,(x2-x1)//2), max(1,(y2-y1)//2))
    c.create_rectangle(x1+r,y1,x2-r,y2, fill=fill,outline='',**kw)
    c.create_rectangle(x1,y1+r,x2,y2-r, fill=fill,outline='',**kw)
    for cx,cy,st in [(x1,y1,90),(x2-2*r,y1,0),(x1,y2-2*r,180),(x2-2*r,y2-2*r,270)]:
        c.create_arc(cx,cy,cx+2*r,cy+2*r,start=st,extent=90,fill=fill,outline=fill,**kw)
    c.create_line(x1+r,y1,x2-r,y1,fill=outline,**kw)
    c.create_line(x1+r,y2,x2-r,y2,fill=outline,**kw)
    c.create_line(x1,y1+r,x1,y2-r,fill=outline,**kw)
    c.create_line(x2,y1+r,x2,y2-r,fill=outline,**kw)
    for cx,cy,st in [(x1,y1,90),(x2-2*r,y1,0),(x1,y2-2*r,180),(x2-2*r,y2-2*r,270)]:
        c.create_arc(cx,cy,cx+2*r,cy+2*r,start=st,extent=90,fill='',outline=outline,style='arc',**kw)

class RoundedCard(tk.Canvas):
    """Canvas card with rounded corners.
    fill_mode=False → body height drives canvas height (for KPI/form cards).
    fill_mode=True  → canvas height drives body height (for table/fill cards).
    """
    def __init__(self, parent, radius=8, card_bg=None, border_color=None,
                 padx=16, pady=12, fill_mode=False, **kw):
        card_bg      = card_bg or THEME.get('card_bg','#FFFFFF')
        border_color = border_color or THEME.get('card_border','#E5E7EB')
        try: par_bg = parent.cget('bg')
        except: par_bg = THEME.get('ct_bg','#F9FAFB')
        super().__init__(parent, bg=par_bg, highlightthickness=0, bd=0, **kw)
        self._r=radius; self._fill=card_bg; self._bc=border_color; self._fm=fill_mode
        self._body=tk.Frame(self, bg=card_bg, padx=padx, pady=pady)
        self._fid=self.create_window(radius, radius, anchor='nw', window=self._body)
        self.bind('<Configure>', self._on_cv)
        if not fill_mode:
            self._body.bind('<Configure>', self._on_body)
    @property
    def body(self): return self._body
    def _on_cv(self, e):
        r=self._r; bw=max(1,e.width-2*r)
        self.itemconfig(self._fid, width=bw)
        if self._fm: self.itemconfig(self._fid, height=max(1,e.height-2*r))
        self.delete('rc')
        if e.width>3 and e.height>3:
            _draw_rr(self,1,1,e.width-1,e.height-1,r,self._fill,self._bc)
            self.tag_lower('rc')
    def _on_body(self, e):
        r=self._r; need_h=e.height+2*r
        if abs(self.winfo_height()-need_h)>1: self.configure(height=need_h)

# ── Rounded button widget ─────────────────────────────────────────────────
class RoundedButton(tk.Canvas):
    """Canvas button with rounded corners. Always draws correct color on resize."""
    _R        = 6
    _DISABLED = "#9CA3AF"

    def __init__(self, parent, text, cmd, bg,
                 icon="", font_size=10, btn_pady=9, **kw):
        kw.pop('padx', None); kw.pop('pady', None)
        self._t   = (f"{icon}  {text}") if icon else text
        self._cmd = cmd
        self._bg  = bg
        self._fnt = (FONT, font_size, 'bold')
        self._on  = False
        try:    par_bg = parent.cget('bg')
        except: par_bg = THEME.get('ct_bg','#F9FAFB')
        import tkinter.font as _tf
        _f = _tf.Font(family=FONT, size=font_size, weight='bold')
        h  = _f.metrics('linespace') + btn_pady * 2 + 2
        # Auto minimum width from text measurement unless caller provides one
        if 'width' not in kw:
            kw['width'] = _f.measure(self._t) + 28
        super().__init__(parent, height=h, highlightthickness=0,
                         bd=0, bg=par_bg, cursor='arrow', **kw)
        self.bind('<Configure>',       self._on_cfg)
        self.bind('<Enter>',           self._hover_in)
        self.bind('<Leave>',           self._hover_out)
        self.bind('<ButtonRelease-1>', self._click)

    @staticmethod
    def _dk(c):
        r,g,b = int(c[1:3],16),int(c[3:5],16),int(c[5:7],16)
        return "#{:02x}{:02x}{:02x}".format(
            max(0,int(r*.82)),max(0,int(g*.82)),max(0,int(b*.82)))

    def _draw(self, color=None):
        if color is None:
            color = self._bg if self._on else self._DISABLED
        self.delete('all')
        w, h = self.winfo_width(), self.winfo_height()
        if w < 4 or h < 4:
            return
        r = min(self._R, w//2, h//2)
        self.create_rectangle(r, 0, w-r, h,   fill=color, outline='')
        self.create_rectangle(0, r, w,   h-r, fill=color, outline='')
        for cx,cy,st in [(0,0,90),(w-2*r,0,0),(0,h-2*r,180),(w-2*r,h-2*r,270)]:
            self.create_arc(cx,cy,cx+2*r,cy+2*r,
                            start=st, extent=90, fill=color, outline=color)
        self.create_text(w//2, h//2, text=self._t,
                         font=self._fnt, fill='#ffffff', anchor='center')

    def _on_cfg(self, e=None):
        # Use event width when available (more reliable than winfo_width before mapping)
        w = getattr(e, 'width', 0) or self.winfo_width()
        if w > 4:
            self._draw()
        else:
            self.after(30, self._draw)

    def _hover_in(self, _=None):
        if self._on: self._draw(self._dk(self._bg))
    def _hover_out(self, _=None): self._draw()
    def _click(self, _=None):
        if self._on and self._cmd:
            self._draw(self._dk(self._dk(self._bg)))
            self.after(120, self._draw)
            self._cmd()

    def enable(self):
        self._on = True
        self.config(cursor='hand2')
        self.after_idle(self._draw)

    def disable(self):
        self._on = False
        self.config(cursor='arrow')
        self.after_idle(self._draw)


def _btn(parent, text, cmd, bg, icon="", font_size=10, btn_pady=9, **kw):
    """Rounded button — always enabled."""
    kw.pop('padx', None); kw.pop('pady', None)
    b = RoundedButton(parent, text, cmd, bg, icon=icon,
                      font_size=font_size, btn_pady=btn_pady, **kw)
    b.enable()
    return b






def _setup_styles():
    s=ttk.Style()
    try: s.theme_use('clam')
    except: pass
    s.configure("POS.Treeview",background="#fff",foreground="#111827",
        fieldbackground="#fff",rowheight=34,font=(FONT,10),borderwidth=0)
    s.configure("POS.Treeview.Heading",background="#111827",foreground="#fff",
        font=(FONT,9,'bold'),relief='flat',padding=(8,6))
    s.map("POS.Treeview",background=[('selected','#2563EB')],foreground=[('selected','#fff')])

def _center(win,w,h):
    win.update_idletasks(); sw,sh=win.winfo_screenwidth(),win.winfo_screenheight()
    win.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")

class ReportsModule:
    def __init__(self,parent,db_manager,current_user=None):
        self.parent=parent; self.db=db_manager
        self.current_user=current_user or {}
        self._rol=(self.current_user.get('rol') or '').lower()
        self._is_admin=self._rol=='administrador'
        self.sales_data=[]; _setup_styles()
        self._build()
        self._quick_filter('today')

    def _stat_card(self,parent,title,value,accent,icon=""):
        outer=RoundedCard(parent,padx=16,pady=14)
        inner=outer.body
        tk.Label(inner,text=f"{icon}  {title}" if icon else title,
                 font=(FONT,9),bg=THEME["card_bg"],fg=THEME["txt_secondary"]).pack(anchor='w')
        lbl=tk.Label(inner,text=value,font=(FONT,18,'bold'),bg=THEME["card_bg"],fg=accent)
        lbl.pack(anchor='w',pady=(4,0))
        tk.Frame(inner,bg=accent,height=3).pack(fill='x',pady=(10,0))
        inner.value_label=lbl; outer.value_label=lbl
        return outer

    def _build(self):
        bg=THEME["ct_bg"]

        # ── Scroll container (para pantallas pequeñas) ─────────────
        scroll_cv = tk.Canvas(self.parent, bg=bg, highlightthickness=0)
        scroll_sb = ttk.Scrollbar(self.parent, orient='vertical',
                                  command=scroll_cv.yview)
        scroll_cv.configure(yscrollcommand=scroll_sb.set)
        scroll_sb.pack(side='right', fill='y')
        scroll_cv.pack(side='left', fill='both', expand=True)

        page = tk.Frame(scroll_cv, bg=bg)
        win_id = scroll_cv.create_window((0, 0), window=page, anchor='nw')

        def _on_cv_resize(e):
            scroll_cv.itemconfig(win_id, width=e.width)
        def _on_page_resize(e):
            scroll_cv.configure(scrollregion=scroll_cv.bbox('all'))
        scroll_cv.bind('<Configure>', _on_cv_resize)
        page.bind('<Configure>', _on_page_resize)

        def _mwheel(e):
            try: scroll_cv.yview_scroll(-1*(e.delta//120), 'units')
            except Exception: pass
        scroll_cv.bind('<Enter>', lambda e: scroll_cv.bind_all('<MouseWheel>', _mwheel))
        scroll_cv.bind('<Leave>', lambda e: scroll_cv.unbind_all('<MouseWheel>'))

        # Alias: todo el contenido va en 'page' en vez de self.parent
        P = page

        # Header
        hdr=tk.Frame(P,bg=bg); hdr.pack(fill='x',padx=28,pady=(20,0))
        tk.Label(hdr,text="📊  Reportes y Estadísticas",font=(FONT,16,'bold'),
                 bg=bg,fg=THEME["txt_primary"]).pack(side='left')
        tk.Frame(P,bg=THEME["card_border"],height=1).pack(fill='x',padx=28,pady=(10,14))

        # Filtros rápidos
        flt_card_outer=RoundedCard(P,padx=16,pady=12)
        flt_card_outer.pack(fill='x',padx=24,pady=(0,14))
        flt_card=flt_card_outer.body
        tk.Label(flt_card,text="Período:",font=(FONT,10,'bold'),
                 bg=THEME["card_bg"],fg=THEME["txt_secondary"]).pack(side='left',padx=(0,12))
        for val,lbl in [('today','Hoy'),('week','Esta semana'),('month','Este mes'),('year','Este año')]:
            b=tk.Label(flt_card,text=lbl,font=(FONT,9,'bold'),cursor='hand2',
                       padx=12,pady=6,bg=THEME["input_brd"],fg=THEME["txt_secondary"])
            b.pack(side='left',padx=2)
            b.bind('<Enter>',lambda e,w=b:w.config(bg=THEME["acc_blue"],fg="#fff"))
            b.bind('<Leave>',lambda e,w=b:w.config(bg=THEME["input_brd"],fg=THEME["txt_secondary"]))
            b.bind('<Button-1>',lambda e,v=val:self._quick_filter(v))

        # Rango personalizado
        tk.Frame(flt_card,bg=THEME["card_border"],width=1).pack(side='left',fill='y',padx=12,pady=2)
        tk.Label(flt_card,text="Desde:",font=(FONT,9,'bold'),bg=THEME["card_bg"],
                 fg=THEME["txt_secondary"]).pack(side='left',padx=(0,6))
        self.date_from=tk.Entry(flt_card,font=(FONT,10),width=11,
                                bg=THEME["input_bg"],fg=THEME["txt_primary"],
                                relief='solid',bd=1,insertbackground=THEME["acc_blue"])
        self.date_from.pack(side='left',padx=(0,8))
        tk.Label(flt_card,text="Hasta:",font=(FONT,9,'bold'),bg=THEME["card_bg"],
                 fg=THEME["txt_secondary"]).pack(side='left',padx=(0,6))
        self.date_to=tk.Entry(flt_card,font=(FONT,10),width=11,
                              bg=THEME["input_bg"],fg=THEME["txt_primary"],
                              relief='solid',bd=1,insertbackground=THEME["acc_blue"])
        self.date_to.pack(side='left',padx=(0,10))
        _btn(flt_card,"Buscar",self.load_report,THEME["acc_blue"],"🔍").pack(side='left')

        # ── Sección Reporte por Rol (solo Administrador) ──────────
        if self._is_admin:
            roles_card_outer=RoundedCard(P,padx=16,pady=10)
            roles_card_outer.pack(fill='x',padx=24,pady=(0,14))
            roles_card=roles_card_outer.body
            # Fila superior: título + selector
            roles_hdr=tk.Frame(roles_card,bg=THEME["card_bg"]); roles_hdr.pack(fill='x',pady=(0,8))
            tk.Label(roles_hdr,text="👤  Filtrar por Usuario / Rol:",font=(FONT,10,'bold'),
                     bg=THEME["card_bg"],fg=THEME["txt_primary"]).pack(side='left')
            _btn(roles_hdr,"Ver Reporte por Rol",self._show_role_report,
                 THEME["acc_purple"],"📊").pack(side='right')
            # Fila de filtros
            roles_flt=tk.Frame(roles_card,bg=THEME["card_bg"]); roles_flt.pack(fill='x')
            # Filtro por rol
            tk.Label(roles_flt,text="Rol:",font=(FONT,9,'bold'),
                     bg=THEME["card_bg"],fg=THEME["txt_secondary"]).pack(side='left',padx=(0,6))
            self.rol_filter_var=tk.StringVar(value="Todos los roles")
            rol_cb=ttk.Combobox(roles_flt,textvariable=self.rol_filter_var,
                values=["Todos los roles","Administrador","Supervisor","Cajero"],
                state='readonly',font=(FONT,9),width=18)
            rol_cb.pack(side='left',padx=(0,16))
            # Filtro por usuario
            tk.Label(roles_flt,text="Usuario:",font=(FONT,9,'bold'),
                     bg=THEME["card_bg"],fg=THEME["txt_secondary"]).pack(side='left',padx=(0,6))
            self.user_filter_var=tk.StringVar(value="Todos los usuarios")
            users=[]
            try: users=self.db.get_all_users()
            except: pass
            user_names=["Todos los usuarios"]+[f"{u['nombre']} ({u['usuario']})" for u in users if u.get('activo',1)]
            self._users_list=users
            self.user_cb=ttk.Combobox(roles_flt,textvariable=self.user_filter_var,
                values=user_names,state='readonly',font=(FONT,9),width=24)
            self.user_cb.pack(side='left',padx=(0,16))
            # Al cambiar rol, filtrar usuarios
            def _filter_users_by_rol(*_):
                sel_rol=self.rol_filter_var.get()
                if sel_rol=="Todos los roles":
                    filtered=self._users_list
                else:
                    filtered=[u for u in self._users_list
                               if u.get('rol','').lower()==sel_rol.lower() and u.get('activo',1)]
                names=["Todos los usuarios"]+[f"{u['nombre']} ({u['usuario']})" for u in filtered]
                self.user_cb['values']=names
                self.user_filter_var.set("Todos los usuarios")
            rol_cb.bind('<<ComboboxSelected>>',_filter_users_by_rol)

        # Cards KPI
        kpi_row=tk.Frame(P,bg=bg); kpi_row.pack(fill='x',padx=24,pady=(0,14))
        self.kpi_count  =self._stat_card(kpi_row,"Total Ventas","0",THEME["acc_blue"],"🛒")
        self.kpi_revenue=self._stat_card(kpi_row,"Ingresos Totales","Gs. 0",THEME["acc_green"],"💰")
        self.kpi_avg    =self._stat_card(kpi_row,"Ticket Promedio","Gs. 0",THEME["acc_amber"],"📈")
        self.kpi_items  =self._stat_card(kpi_row,"Productos Vendidos","0",THEME["acc_purple"],"📦")
        for i,w in enumerate([self.kpi_count,self.kpi_revenue,self.kpi_avg,self.kpi_items]):
            w.grid(row=0,column=i,sticky='nsew',padx=4); kpi_row.columnconfigure(i,weight=1)

        # Tabla ventas
        tbl_outer=RoundedCard(P,padx=0,pady=0,fill_mode=False)
        tbl_outer.pack(fill='x',padx=24,pady=(0,10))
        tbl=tbl_outer.body

        tbl_hdr=tk.Frame(tbl,bg=THEME["card_bg"],padx=16,pady=10); tbl_hdr.pack(fill='x')
        tk.Label(tbl_hdr,text="Detalle de Ventas",font=(FONT,11,'bold'),
                 bg=THEME["card_bg"],fg=THEME["txt_primary"]).pack(side='left')
        self.lbl_count=tk.Label(tbl_hdr,text="",font=(FONT,9),
                                bg=THEME["card_bg"],fg=THEME["txt_secondary"]); self.lbl_count.pack(side='left',padx=8)
        _btn(tbl_hdr,"PDF",self.export_pdf,THEME["acc_rose"],"📄").pack(side='right',padx=(4,0))
        _btn(tbl_hdr,"Excel",self.export_excel,THEME["acc_green"],"📊").pack(side='right',padx=(4,0))
        _btn(tbl_hdr,"CSV",self.export_csv,THEME["acc_cyan"],"📥").pack(side='right',padx=(0,0))
        tk.Frame(tbl,bg=THEME["card_border"],height=1).pack(fill='x')

        cols=('ID','Fecha','Cliente','Total','Método','Usuario','Estado')
        self.tree=ttk.Treeview(tbl,columns=cols,show='headings',style='POS.Treeview',height=10)
        widths={'ID':50,'Fecha':145,'Cliente':170,'Total':110,'Método':110,'Usuario':120,'Estado':75}
        for col in cols:
            self.tree.heading(col,text=col)
            self.tree.column(col,width=widths[col],anchor='w' if col in('Fecha','Cliente') else 'center')
        sb=ttk.Scrollbar(tbl,orient='vertical',command=self.tree.yview)
        self.tree.configure(yscrollcommand=sb.set)
        sb.pack(side='right',fill='y',padx=(0,4),pady=4); self.tree.pack(fill='both',expand=True,padx=4,pady=4)
        self.tree.bind('<Double-Button-1>',self._show_sale_detail)
        self.tree.tag_configure('odd',background=THEME["row_odd"])
        self.tree.tag_configure('even',background=THEME["row_even"])
        self.tree.tag_configure('credit',foreground=THEME["acc_purple"])

        # ── Gráfico de ventas por día ──────────────────────────────────────────
        chart_outer=RoundedCard(P,padx=16,pady=12)
        chart_outer.pack(fill='x',padx=24,pady=(0,10))
        chart_card=chart_outer.body
        chart_hdr=tk.Frame(chart_card,bg=THEME["card_bg"]); chart_hdr.pack(fill='x',pady=(0,8))
        tk.Label(chart_hdr,text="📊  Ventas por Día",font=(FONT,11,'bold'),
                 bg=THEME["card_bg"],fg=THEME["txt_primary"]).pack(side='left')
        self._bar_chart = BarChart(chart_card, height=160)
        self._bar_chart.pack(fill='x', padx=4, pady=(0,4))

    def _quick_filter(self,period):
        today=datetime.now()
        if period=='today':
            d_from=d_to=today.strftime('%Y-%m-%d')
        elif period=='week':
            d_from=(today-timedelta(days=today.weekday())).strftime('%Y-%m-%d')
            d_to=today.strftime('%Y-%m-%d')
        elif period=='month':
            d_from=today.strftime('%Y-%m-01')
            d_to=today.strftime('%Y-%m-%d')
        else:
            d_from=today.strftime('%Y-01-01')
            d_to=today.strftime('%Y-%m-%d')
        self.date_from.delete(0,'end'); self.date_from.insert(0,d_from)
        self.date_to.delete(0,'end'); self.date_to.insert(0,d_to)
        self.load_report()

    def load_report(self):
        d_from=self.date_from.get().strip(); d_to=self.date_to.get().strip()
        if not d_from or not d_to: return
        try:
            datetime.strptime(d_from,'%Y-%m-%d'); datetime.strptime(d_to,'%Y-%m-%d')
        except ValueError: messagebox.showerror("Error","Use formato YYYY-MM-DD"); return

        for i in self.tree.get_children(): self.tree.delete(i)
        self.sales_data=self.db.get_sales_by_date(d_from,d_to)
        summary=self.db.get_sales_summary(d_from,d_to)

        self.kpi_count.value_label.config(text=str(summary.get('total_sales',0)))
        self.kpi_revenue.value_label.config(text=format_currency(summary.get('total_revenue',0)))
        self.kpi_avg.value_label.config(text=format_currency(summary.get('average_sale',0)))

        # Productos vendidos
        try:
            items=self.db.execute_scalar(
                """SELECT COALESCE(SUM(d.cantidad),0) FROM detalle_ventas d
                   JOIN ventas v ON d.venta_id=v.id
                   WHERE v.fecha BETWEEN ? AND ?""",(d_from,d_to+' 23:59:59')) or 0
            self.kpi_items.value_label.config(text=str(int(items)))
        except: pass

        for idx,s in enumerate(self.sales_data):
            tag_row='odd' if idx%2 else 'even'
            tags=(tag_row,'credit') if s.get('payment_method','')=='Crédito' else (tag_row,)
            cajero=s.get('cajero_nombre') or s.get('cashier_name','') or '—'
            self.tree.insert('','end',tags=tags,values=(
                s.get('id',''),
                s.get('sale_date','')[:16] if s.get('sale_date') else '',
                s.get('customer_name','') or 'Consumidor Final',
                format_currency(s.get('total',0)),
                s.get('payment_method',''),
                cajero,
                'Crédito' if s.get('payment_method')=='Crédito' else 'Contado'))
        self.lbl_count.config(text=f"({len(self.sales_data)} registros)")

        # ── Actualizar gráfico de barras ──────────────────────────────────────
        try:
            from datetime import datetime as _dt, timedelta as _td
            dt_from = _dt.strptime(d_from, '%Y-%m-%d')
            dt_to   = _dt.strptime(d_to,   '%Y-%m-%d')
            delta   = (dt_to - dt_from).days + 1
            # Agrupar totales por día
            daily = {}
            for s in self.sales_data:
                day = str(s.get('sale_date',''))[:10]
                daily[day] = daily.get(day, 0) + s.get('total', 0)
            if delta <= 14:
                chart_data = []
                for i in range(delta):
                    day = (dt_from + _td(days=i)).strftime('%Y-%m-%d')
                    lbl = (dt_from + _td(days=i)).strftime('%d/%m')
                    chart_data.append((lbl, daily.get(day, 0)))
            else:
                # Más de 14 días: mostrar los últimos 14 días con datos
                items = sorted(daily.items())[-14:]
                chart_data = [(_dt.strptime(d,'%Y-%m-%d').strftime('%d/%m'), v)
                              for d,v in items]
            self._bar_chart.set_data(chart_data)
        except Exception:
            pass

    def _show_sale_detail(self,_=None):
        sel=self.tree.selection()
        if not sel: return
        sale_id=self.tree.item(sel[0])['values'][0]
        details=self.db.get_sale_details(sale_id)
        if not details: return
        sale=next((s for s in self.sales_data if s.get('id')==sale_id),{})
        dlg=tk.Toplevel(self.parent)
        _set_icon(dlg); dlg.title(f"Venta #{sale_id}")
        dlg.configure(bg=THEME["ct_bg"]); dlg.transient(self.parent); _center(dlg,580,460)
        hdr=tk.Frame(dlg,bg=THEME["sb_bg"],pady=12); hdr.pack(fill='x')
        tk.Label(hdr,text=f"🧾  Venta #{sale_id} — {sale.get('customer_name','') or 'Consumidor Final'}",
                 font=(FONT,12,'bold'),bg=THEME["sb_bg"],fg="#fff").pack(anchor='w',padx=16)
        tk.Label(hdr,text=f"Total: {format_currency(sale.get('total',0))}  ·  {sale.get('sale_date','')[:16]}",
                 font=(FONT,9),bg=THEME["sb_bg"],fg="#94A3B8").pack(anchor='w',padx=16)
        tbl=tk.Frame(dlg,bg=THEME["card_bg"]); tbl.pack(fill='both',expand=True,padx=16,pady=16)
        cols2=('Producto','Cantidad','Precio Unit.','Subtotal')
        tree2=ttk.Treeview(tbl,columns=cols2,show='headings',height=10,style='POS.Treeview')
        for col,w in zip(cols2,[240,80,120,120]):
            tree2.heading(col,text=col); tree2.column(col,width=w,anchor='w' if col=='Producto' else 'center')
        sb2=ttk.Scrollbar(tbl,orient='vertical',command=tree2.yview)
        tree2.configure(yscrollcommand=sb2.set)
        sb2.pack(side='right',fill='y'); tree2.pack(fill='both',expand=True)
        for i,d in enumerate(details):
            tree2.insert('','end',tags=('odd' if i%2 else 'even',),values=(
                d.get('product_name',''),d.get('quantity',0),
                format_currency(d.get('unit_price',0)),
                format_currency(d.get('quantity',0)*d.get('unit_price',0))))
        tree2.tag_configure('odd',background=THEME["row_odd"]); tree2.tag_configure('even',background=THEME["row_even"])
        _btn(dlg,"Cerrar",dlg.destroy,THEME["btn_secondary"]).pack(pady=10)

    def _show_role_report(self):
        """Ventana de reporte detallado por rol/usuario con KPIs y tabla."""
        d_from=self.date_from.get().strip(); d_to=self.date_to.get().strip()
        if not d_from or not d_to:
            messagebox.showinfo("Info","Seleccione un rango de fechas primero"); return

        # Obtener todas las ventas del período
        all_sales=self.db.get_sales_by_date(d_from,d_to)

        # Aplicar filtros
        sel_rol=self.rol_filter_var.get()
        sel_user=self.user_filter_var.get()

        filtered=[]
        for s in all_sales:
            cajero=s.get('cajero_nombre') or s.get('cashier_name','')
            # Filtro por nombre de usuario
            if sel_user!="Todos los usuarios":
                u_nombre=sel_user.split(' (')[0]
                if cajero!=u_nombre: continue
            elif sel_rol!="Todos los roles":
                # Filtrar por rol: buscar en lista de usuarios
                matched_names=[u['nombre'] for u in self._users_list
                               if u.get('rol','').lower()==sel_rol.lower()]
                if cajero not in matched_names: continue
            filtered.append(s)

        # Agrupar por usuario
        from collections import defaultdict
        by_user=defaultdict(list)
        for s in filtered:
            cajero=s.get('cajero_nombre') or s.get('cashier_name','') or 'Sin asignar'
            by_user[cajero].append(s)

        # Ventana
        dlg=tk.Toplevel(self.parent)
        _set_icon(dlg); dlg.title("Reporte por Rol/Usuario")
        dlg.configure(bg=THEME["ct_bg"]); dlg.resizable(True,True)
        dlg.transient(self.parent); _center(dlg,860,620); dlg.minsize(700,500)

        # Botones PRIMERO
        btn_bar=tk.Frame(dlg,bg=THEME["card_bg"],padx=16,pady=10)
        btn_bar.pack(fill='x',side='bottom')
        tk.Frame(dlg,bg=THEME["card_border"],height=1).pack(fill='x',side='bottom')

        # Header
        hdr=tk.Frame(dlg,bg=THEME["sb_bg"],pady=12); hdr.pack(fill='x')
        filtro_txt=(f"Rol: {sel_rol}" if sel_rol!="Todos los roles" else "") +                    ("  |  " if sel_rol!="Todos los roles" and sel_user!="Todos los usuarios" else "") +                    (f"Usuario: {sel_user.split(' (')[0]}" if sel_user!="Todos los usuarios" else "")
        filtro_txt=filtro_txt or "Todos los usuarios"
        tk.Label(hdr,text=f"📊  Reporte por Rol/Usuario",
                 font=(FONT,13,'bold'),bg=THEME["sb_bg"],fg="#fff").pack(anchor='w',padx=16)
        tk.Label(hdr,text=f"Período: {d_from}  →  {d_to}   |   Filtro: {filtro_txt}",
                 font=(FONT,9),bg=THEME["sb_bg"],fg="#8AABD4").pack(anchor='w',padx=18)

        # Canvas scroll
        canvas=tk.Canvas(dlg,bg=THEME["ct_bg"],highlightthickness=0)
        sbar=tk.Scrollbar(dlg,orient='vertical',command=canvas.yview)
        canvas.configure(yscrollcommand=sbar.set)
        sbar.pack(side='right',fill='y'); canvas.pack(fill='both',expand=True)
        body=tk.Frame(canvas,bg=THEME["ct_bg"],padx=20,pady=16)
        canvas.create_window((0,0),window=body,anchor='nw',tags='body')
        body.bind('<Configure>',lambda e:canvas.configure(scrollregion=canvas.bbox('all')))
        canvas.bind('<Configure>',lambda e:canvas.itemconfig('body',width=e.width))
        def _mw(e):
            try: canvas.yview_scroll(-1*(e.delta//120),'units')
            except: canvas.unbind_all('<MouseWheel>')
        canvas.bind('<Enter>',lambda e:canvas.bind_all('<MouseWheel>',_mw))
        canvas.bind('<Leave>',lambda e:canvas.unbind_all('<MouseWheel>'))

        # ── Resumen global ──
        total_global=sum(s.get('total',0) for s in filtered)
        sec0=tk.Frame(body,bg=THEME["card_border"]); sec0.pack(fill='x',pady=(0,14))
        inn0=tk.Frame(sec0,bg=THEME["card_bg"],padx=16,pady=12); inn0.pack(fill='x',padx=1,pady=1)
        tk.Label(inn0,text="Resumen Global del Período",font=(FONT,10,'bold'),
                 bg=THEME["card_bg"],fg=THEME["acc_blue"]).pack(anchor='w',pady=(0,8))
        kpi_row=tk.Frame(inn0,bg=THEME["card_bg"]); kpi_row.pack(fill='x')
        for i,(lbl,val,col) in enumerate([
            ("Total Ventas",str(len(filtered)),THEME["acc_blue"]),
            ("Ingresos",format_currency(total_global),THEME["acc_green"]),
            ("Usuarios",str(len(by_user)),THEME["acc_purple"]),
            ("Promedio/Venta",format_currency(total_global/len(filtered)) if filtered else "—",THEME["acc_amber"]),
        ]):
            card=tk.Frame(kpi_row,bg=THEME["card_border"]); card.grid(row=0,column=i,sticky='nsew',padx=4)
            inn=tk.Frame(card,bg=THEME["card_bg"],padx=12,pady=10); inn.pack(fill='both',padx=1,pady=1)
            tk.Label(inn,text=lbl,font=(FONT,8),bg=THEME["card_bg"],fg=THEME["txt_secondary"]).pack(anchor='w')
            tk.Label(inn,text=val,font=(FONT,13,'bold'),bg=THEME["card_bg"],fg=col).pack(anchor='w')
            kpi_row.columnconfigure(i,weight=1)

        # ── Tarjeta por usuario ──
        for cajero_name,sales in sorted(by_user.items()):
            # Buscar rol del usuario
            u_rol='—'
            for u in self._users_list:
                if u.get('nombre')==cajero_name:
                    u_rol=u.get('rol','—').title(); break
            total_u=sum(s.get('total',0) for s in sales)
            contado=[s for s in sales if s.get('payment_method','')!='Crédito']
            credito=[s for s in sales if s.get('payment_method','')=='Crédito']

            sec=tk.Frame(body,bg=THEME["card_border"]); sec.pack(fill='x',pady=(0,12))
            inn=tk.Frame(sec,bg=THEME["card_bg"],padx=16,pady=12); inn.pack(fill='x',padx=1,pady=1)

            # Cabecera usuario
            uh=tk.Frame(inn,bg=THEME["card_bg"]); uh.pack(fill='x',pady=(0,10))
            ROL_COLORS={'administrador':THEME["acc_blue"],'supervisor':THEME["acc_amber"],
                        'cajero':THEME["acc_green"]}
            rol_col=ROL_COLORS.get(u_rol.lower(),THEME["acc_purple"])
            tk.Label(uh,text=f"👤  {cajero_name}",font=(FONT,11,'bold'),
                     bg=THEME["card_bg"],fg=THEME["txt_primary"]).pack(side='left')
            tk.Label(uh,text=f"  {u_rol}",font=(FONT,9,'bold'),
                     bg=rol_col,fg="#fff",padx=8,pady=2).pack(side='left',padx=8)

            # KPIs del usuario
            kpi_u=tk.Frame(inn,bg=THEME["card_bg"]); kpi_u.pack(fill='x',pady=(0,10))
            for i,(lbl,val,col) in enumerate([
                ("Ventas totales",str(len(sales)),THEME["acc_blue"]),
                ("Ingresos",format_currency(total_u),THEME["acc_green"]),
                ("Ventas contado",str(len(contado)),THEME["acc_amber"]),
                ("Ventas crédito",str(len(credito)),THEME["acc_purple"]),
                ("Ticket promedio",format_currency(total_u/len(sales)) if sales else "—",THEME["txt_secondary"]),
            ]):
                c2=tk.Frame(kpi_u,bg=THEME["card_border"]); c2.grid(row=0,column=i,sticky='nsew',padx=3)
                i2=tk.Frame(c2,bg=THEME["card_bg"],padx=10,pady=8); i2.pack(fill='both',padx=1,pady=1)
                tk.Label(i2,text=lbl,font=(FONT,7),bg=THEME["card_bg"],fg=THEME["txt_secondary"]).pack(anchor='w')
                tk.Label(i2,text=val,font=(FONT,11,'bold'),bg=THEME["card_bg"],fg=col).pack(anchor='w')
                kpi_u.columnconfigure(i,weight=1)

            # Tabla de ventas del usuario
            tk.Label(inn,text="Detalle de ventas:",font=(FONT,9,'bold'),
                     bg=THEME["card_bg"],fg=THEME["txt_secondary"]).pack(anchor='w',pady=(0,4))
            u_cols=('ID','Fecha','Cliente','Total','Método')
            u_tree=ttk.Treeview(inn,columns=u_cols,show='headings',
                                style='POS.Treeview',height=min(len(sales),5))
            for col,w in zip(u_cols,[50,145,180,110,110]):
                u_tree.heading(col,text=col)
                u_tree.column(col,width=w,anchor='w' if col in('Fecha','Cliente') else 'center')
            for si,s in enumerate(sorted(sales,key=lambda x:x.get('sale_date',''),reverse=True)):
                tag='odd' if si%2 else 'even'
                u_tree.insert('','end',tags=(tag,),values=(
                    s.get('id',''),
                    (s.get('sale_date','') or '')[:16],
                    s.get('customer_name','') or 'Consumidor Final',
                    format_currency(s.get('total',0)),
                    s.get('payment_method',''),
                ))
            u_tree.tag_configure('odd',background=THEME["row_odd"])
            u_tree.tag_configure('even',background=THEME["row_even"])
            u_tree.pack(fill='x')

        if not by_user:
            tk.Label(body,text="No hay ventas registradas para este filtro en el período seleccionado.",
                     font=(FONT,10),bg=THEME["ct_bg"],fg=THEME["txt_secondary"]).pack(pady=40)

        # Botón exportar + cerrar
        def _exp_role():
            self._export_role_report(by_user,d_from,d_to,filtro_txt)
        b1=_btn(btn_bar,"Exportar Excel",_exp_role,THEME["acc_green"],"📊"); b1.pack(side='left',padx=(0,4),fill='x',expand=True)
        b2=_btn(btn_bar,"Cerrar",dlg.destroy,THEME["btn_secondary"],"✕"); b2.pack(side='left',fill='x',expand=True)

    def _export_role_report(self,by_user,d_from,d_to,filtro_txt):
        """Exporta el reporte por roles a Excel."""
        try:
            import openpyxl
            from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
            from openpyxl.utils import get_column_letter
            from datetime import datetime as _dt
            try: from utils.company_header import get_company as _get_company
            except: _get_company=lambda:{}
            co=_get_company()
            empresa=co.get("razon_social","Venialgo POS")
            thin=Side(style='thin',color='D1D5DB')
            brd=Border(left=thin,right=thin,top=thin,bottom=thin)
            def hdr_s(cell,bg="111827",fg="FFFFFF"):
                cell.font=Font(name='Arial',bold=True,color=fg,size=10)
                cell.fill=PatternFill('solid',start_color=bg)
                cell.alignment=Alignment(horizontal='center',vertical='center')
                cell.border=brd
            wb=openpyxl.Workbook()
            # ── Hoja resumen por usuario ──
            ws=wb.active; ws.title="Resumen por Usuario"
            ws.merge_cells("A1:G1"); ws["A1"]=empresa
            ws["A1"].font=Font(name='Arial',bold=True,size=14,color="2563EB")
            ws["A1"].alignment=Alignment(horizontal='center')
            ws.row_dimensions[1].height=24
            ws.merge_cells("A2:G2")
            ws["A2"]=f"REPORTE POR ROL/USUARIO  |  Período: {d_from} → {d_to}  |  Filtro: {filtro_txt}"
            ws["A2"].font=Font(name='Arial',size=9,color="6B7280",italic=True)
            ws["A2"].alignment=Alignment(horizontal='center')
            ws.row_dimensions[2].height=14
            ws.merge_cells("A3:G3")
            ws["A3"]=f"Generado: {_dt.now().strftime('%d/%m/%Y %H:%M')}   |   Usuarios: {len(by_user)}"
            ws["A3"].font=Font(name='Arial',size=9,color="6B7280",italic=True)
            ws["A3"].alignment=Alignment(horizontal='center')
            ws.row_dimensions[3].height=14
            for ci,h in enumerate(["Usuario","Rol","Total Ventas","Contado","Crédito","Ingresos","Ticket Prom."],1):
                hdr_s(ws.cell(row=4,column=ci,value=h))
            for ci,w in enumerate([22,14,13,12,12,16,14],1):
                ws.column_dimensions[get_column_letter(ci)].width=w
            ri=5
            for cajero_name,sales in sorted(by_user.items()):
                u_rol='—'
                for u in self._users_list:
                    if u.get('nombre')==cajero_name: u_rol=u.get('rol','—').title(); break
                total_u=sum(s.get('total',0) for s in sales)
                contado=len([s for s in sales if s.get('payment_method','')!='Crédito'])
                credito=len([s for s in sales if s.get('payment_method','')=='Crédito'])
                vals=[cajero_name,u_rol,len(sales),contado,credito,total_u,
                      total_u/len(sales) if sales else 0]
                bg="F0FDF4" if u_rol.lower()=='cajero' else ("EFF6FF" if u_rol.lower()=='administrador' else "FFFBEB")
                for ci,v in enumerate(vals,1):
                    cell=ws.cell(row=ri,column=ci,value=v)
                    cell.font=Font(name='Arial',size=9)
                    cell.fill=PatternFill('solid',start_color=bg)
                    cell.alignment=Alignment(horizontal='center' if ci>1 else 'left',vertical='center')
                    cell.border=brd
                ri+=1
                # Sub-tabla de ventas
                ws2=wb.create_sheet(cajero_name[:28])
                ws2.merge_cells("A1:F1"); ws2["A1"]=f"{empresa} — Ventas de: {cajero_name}"
                ws2["A1"].font=Font(name='Arial',bold=True,size=12,color="2563EB")
                ws2["A1"].alignment=Alignment(horizontal='center')
                for ci2,h2 in enumerate(["ID","Fecha","Cliente","Total","Método","Estado"],1):
                    hdr_s(ws2.cell(row=2,column=ci2,value=h2))
                for ci2,w2 in enumerate([8,18,22,14,14,12],1):
                    ws2.column_dimensions[get_column_letter(ci2)].width=w2
                for si,s in enumerate(sorted(sales,key=lambda x:x.get('sale_date',''),reverse=True),3):
                    bg2="F9FAFB" if si%2==1 else "FFFFFF"
                    for ci2,v2 in enumerate([
                        s.get('id',''),
                        (s.get('sale_date','') or '')[:16],
                        s.get('customer_name','') or 'Consumidor Final',
                        s.get('total',0),
                        s.get('payment_method',''),
                        'Crédito' if s.get('payment_method','')=='Crédito' else 'Contado'
                    ],1):
                        cell=ws2.cell(row=si,column=ci2,value=v2)
                        cell.font=Font(name='Arial',size=9)
                        cell.fill=PatternFill('solid',start_color=bg2)
                        cell.alignment=Alignment(horizontal='center' if ci2 not in(2,3) else 'left',vertical='center')
                        cell.border=brd
            import os,datetime
            docs=os.path.join(os.path.expanduser("~"),"Documents")
            os.makedirs(docs,exist_ok=True)
            ts=datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            path=os.path.join(docs,f"reporte_roles_{ts}.xlsx")
            wb.save(path)
            messagebox.showinfo("✅ Exportado",f"Archivo guardado en:\n{path}")
            os.startfile(os.path.dirname(path))
        except ImportError:
            messagebox.showerror("Error","Se requiere openpyxl: pip install openpyxl")
        except Exception as ex:
            messagebox.showerror("Error al exportar",str(ex))

    def export_excel(self):
        """Exporta el reporte actual a Excel usando openpyxl."""
        if not self.sales_data:
            messagebox.showinfo("Info","Sin datos para exportar"); return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            messagebox.showerror("Error","Se requiere openpyxl:\npip install openpyxl"); return
        try:
            from tkinter import filedialog
            import datetime, os
            d_from=self.date_from.get().strip(); d_to=self.date_to.get().strip()
            path=filedialog.asksaveasfilename(
                defaultextension='.xlsx',
                filetypes=[('Excel','*.xlsx')],
                title="Guardar Excel",
                initialfile=f"ventas_{d_from}_{d_to}.xlsx")
            if not path: return

            thin=Side(style='thin',color='D1D5DB')
            brd=Border(left=thin,right=thin,top=thin,bottom=thin)

            wb=openpyxl.Workbook()
            ws=wb.active; ws.title="Ventas"

            # Título
            try:
                from utils.company_header import get_company as _gc
                empresa=_gc().get("razon_social","Venialgo POS")
            except Exception:
                empresa="Venialgo POS"

            ws.merge_cells("A1:G1"); ws["A1"]=empresa
            ws["A1"].font=Font(name='Arial',bold=True,size=14,color="2563EB")
            ws["A1"].alignment=Alignment(horizontal='center')
            ws.row_dimensions[1].height=22

            ws.merge_cells("A2:G2")
            ws["A2"]=f"Reporte de Ventas  |  Período: {d_from} → {d_to}"
            ws["A2"].font=Font(name='Arial',size=9,color="6B7280",italic=True)
            ws["A2"].alignment=Alignment(horizontal='center')

            # Encabezados
            headers=["ID","Fecha","Cliente","Total","Método","Cajero","Estado"]
            for ci,h in enumerate(headers,1):
                cell=ws.cell(row=3,column=ci,value=h)
                cell.font=Font(name='Arial',bold=True,color="FFFFFF",size=10)
                cell.fill=PatternFill('solid',start_color="111827")
                cell.alignment=Alignment(horizontal='center',vertical='center')
                cell.border=brd

            for ci,w in enumerate([8,18,24,14,14,18,10],1):
                ws.column_dimensions[get_column_letter(ci)].width=w

            for ri,s in enumerate(self.sales_data,4):
                bg="F9FAFB" if ri%2==0 else "FFFFFF"
                cajero=s.get('cajero_nombre') or s.get('cashier_name','') or '—'
                vals=[s.get('id',''),
                      (s.get('sale_date','') or '')[:16],
                      s.get('customer_name','') or 'Consumidor Final',
                      s.get('total',0),
                      s.get('payment_method',''),
                      cajero,
                      'Crédito' if s.get('payment_method','')=='Crédito' else 'Contado']
                for ci,v in enumerate(vals,1):
                    cell=ws.cell(row=ri,column=ci,value=v)
                    cell.font=Font(name='Arial',size=9)
                    cell.fill=PatternFill('solid',start_color=bg)
                    cell.alignment=Alignment(
                        horizontal='center' if ci not in(2,3) else 'left',
                        vertical='center')
                    cell.border=brd

            wb.save(path)
            messagebox.showinfo("✅ Exportado",f"Excel guardado en:\n{path}")
            try: os.startfile(os.path.dirname(path))
            except Exception: pass
        except Exception as ex:
            messagebox.showerror("Error al exportar",str(ex))

    def export_pdf(self):
        """Exporta el reporte como texto formateado (.txt abierto como reporte)."""
        if not self.sales_data:
            messagebox.showinfo("Info","Sin datos para exportar"); return
        from tkinter import filedialog
        import os
        d_from=self.date_from.get().strip(); d_to=self.date_to.get().strip()
        path=filedialog.asksaveasfilename(
            defaultextension='.txt',
            filetypes=[('Texto / PDF','*.txt'),('Todos','*.*')],
            title="Guardar Reporte",
            initialfile=f"reporte_{d_from}_{d_to}.txt")
        if not path: return
        try:
            from datetime import datetime as _dt
            try:
                from utils.company_header import get_company as _gc
                empresa=_gc().get("razon_social","Venialgo POS")
            except Exception:
                empresa="Venialgo POS"

            SEP="═"*56; sep="-"*56
            lines=[
                SEP,
                empresa.center(56),
                "REPORTE DE VENTAS".center(56),
                f"Período: {d_from}  →  {d_to}".center(56),
                f"Generado: {_dt.now().strftime('%d/%m/%Y %H:%M')}".center(56),
                SEP,"",
                f"  {'#':<6} {'Fecha':<18} {'Total':>12}  {'Método'}",
                sep,
            ]
            total_global=0
            for s in self.sales_data:
                lines.append(
                    f"  #{str(s.get('id','')):<5} "
                    f"{str(s.get('sale_date',''))[:16]:<18} "
                    f"{format_currency(s.get('total',0)):>12}  "
                    f"{s.get('payment_method','')}")
                total_global+=s.get('total',0)
            lines+=[sep,
                    f"  {'TOTAL':>25} {format_currency(total_global):>12}",
                    f"  {'Transacciones:':>25} {len(self.sales_data):>12}",
                    SEP,""]
            with open(path,'w',encoding='utf-8') as f:
                f.write('\n'.join(lines))
            messagebox.showinfo("✅ Exportado",f"Reporte guardado en:\n{path}")
            try: os.startfile(path)
            except Exception: pass
        except Exception as ex:
            messagebox.showerror("Error al exportar",str(ex))

    def export_csv(self):
        if not self.sales_data: messagebox.showinfo("Info","Sin datos para exportar"); return
        from tkinter import filedialog
        import csv
        path=filedialog.asksaveasfilename(defaultextension='.csv',
            filetypes=[('CSV','*.csv')],title="Guardar reporte")
        if not path: return
        try:
            with open(path,'w',newline='',encoding='utf-8-sig') as f:
                w=csv.writer(f)
                w.writerow(['ID','Fecha','Cliente','Total','Método'])
                for s in self.sales_data:
                    w.writerow([s.get('id',''),s.get('sale_date',''),
                        s.get('customer_name','') or 'Consumidor Final',
                        s.get('total',0),s.get('payment_method','')])
            messagebox.showinfo("✅ Exportado",f"Archivo guardado:\n{path}")
        except Exception as e: messagebox.showerror("Error",str(e))
