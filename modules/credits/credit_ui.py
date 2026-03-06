"""Módulo Gestión de Créditos — Diseño moderno Venialgo POS"""
import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime, timedelta
from config.settings import COLORS, FONTS as _F, SPACING, BUTTON_STYLES, CREDIT
try:
    from utils.company_header import get_company as _get_company
except ImportError:
    def _get_company(): return {}
from utils.formatters import format_currency, format_date
try:
    from utils.window_icon import set_icon as _set_icon
except ImportError:
    def _set_icon(w): pass

THEME = {"ct_bg":"#F9FAFB","card_bg":"#FFFFFF","card_border":"#E5E7EB","sb_bg":"#111827",
    "txt_primary":"#111827","txt_secondary":"#6B7280","txt_white":"#FFFFFF",
    "acc_blue":"#2563EB","acc_green":"#059669","acc_amber":"#D97706","acc_rose":"#E11D48",
    "acc_purple":"#7C3AED","btn_danger":"#DC2626","btn_secondary":"#6B7280",
    "input_bg":"#FFFFFF","input_brd":"#D1D5DB","input_foc":"#2563EB",
    "row_odd":"#F9FAFB","row_even":"#FFFFFF"}
FONT="Segoe UI"

# ── Rounded card widget ───────────────────────────────────────────────────
def _draw_rr(c, x1, y1, x2, y2, r, fill, outline):
    """Rounded rect on canvas; tag='rc'."""
    kw = {'tags': 'rc'}
    r = min(r, max(1,(x2-x1)//2), max(1,(y2-y1)//2))
    c.create_rectangle(x1+r,y1,x2-r,y2, fill=fill,outline='',**kw)
    c.create_rectangle(x1,y1+r,x2,y2-r, fill=fill,outline='',**kw)
    for cx,cy,st in [(x1,y1,90),(x2-2*r,y1,0),(x1,y2-2*r,180),(x2-2*r,y2-2*r,270)]:
        c.create_arc(cx,cy,cx+2*r,cy+2*r,start=st,extent=90,fill=fill,outline=fill,**kw)
    c.create_line(x1+r,y1,x2-r,y1,fill=outline,**kw)
    c.create_line(x1+r,y2,x2-r,y2,fill=outline,**kw)
    c.create_line(x1,y1+r,x1,y2-r,fill=outline,**kw)
    c.create_line(x2,y1+r,x2,y2-r,fill=outline,**kw)
    for cx,cy,st in [(x1,y1,90),(x2-2*r,y1,0),(x1,y2-2*r,180),(x2-2*r,y2-2*r,270)]:
        c.create_arc(cx,cy,cx+2*r,cy+2*r,start=st,extent=90,fill='',outline=outline,style='arc',**kw)

class RoundedCard(tk.Canvas):
    """Canvas card with rounded corners.
    fill_mode=False → body height drives canvas height (for KPI/form cards).
    fill_mode=True  → canvas height drives body height (for table/fill cards).
    """
    def __init__(self, parent, radius=8, card_bg=None, border_color=None,
                 padx=16, pady=12, fill_mode=False, **kw):
        card_bg      = card_bg or THEME.get('card_bg','#FFFFFF')
        border_color = border_color or THEME.get('card_border','#E5E7EB')
        try: par_bg = parent.cget('bg')
        except: par_bg = THEME.get('ct_bg','#F9FAFB')
        super().__init__(parent, bg=par_bg, highlightthickness=0, bd=0, **kw)
        self._r=radius; self._fill=card_bg; self._bc=border_color; self._fm=fill_mode
        self._body=tk.Frame(self, bg=card_bg, padx=padx, pady=pady)
        self._fid=self.create_window(radius, radius, anchor='nw', window=self._body)
        self.bind('<Configure>', self._on_cv)
        if not fill_mode:
            self._body.bind('<Configure>', self._on_body)
    @property
    def body(self): return self._body
    def _on_cv(self, e):
        r=self._r; bw=max(1,e.width-2*r)
        self.itemconfig(self._fid, width=bw)
        if self._fm: self.itemconfig(self._fid, height=max(1,e.height-2*r))
        self.delete('rc')
        if e.width>3 and e.height>3:
            _draw_rr(self,1,1,e.width-1,e.height-1,r,self._fill,self._bc)
            self.tag_lower('rc')
    def _on_body(self, e):
        r=self._r; need_h=e.height+2*r
        if abs(self.winfo_height()-need_h)>1: self.configure(height=need_h)

# ── Rounded button widget ─────────────────────────────────────────────────
class RoundedButton(tk.Canvas):
    """Canvas button with rounded corners. Always draws correct color on resize."""
    _R        = 6
    _DISABLED = "#9CA3AF"

    def __init__(self, parent, text, cmd, bg,
                 icon="", font_size=10, btn_pady=9, **kw):
        kw.pop('padx', None); kw.pop('pady', None)
        self._t   = (f"{icon}  {text}") if icon else text
        self._cmd = cmd
        self._bg  = bg
        self._fnt = (FONT, font_size, 'bold')
        self._on  = False
        try:    par_bg = parent.cget('bg')
        except: par_bg = THEME.get('ct_bg','#F9FAFB')
        import tkinter.font as _tf
        _f = _tf.Font(family=FONT, size=font_size, weight='bold')
        h  = _f.metrics('linespace') + btn_pady * 2 + 2
        super().__init__(parent, height=h, highlightthickness=0,
                         bd=0, bg=par_bg, cursor='arrow', **kw)
        self.bind('<Configure>',       self._on_cfg)
        self.bind('<Enter>',           self._hover_in)
        self.bind('<Leave>',           self._hover_out)
        self.bind('<ButtonRelease-1>', self._click)

    @staticmethod
    def _dk(c):
        r,g,b = int(c[1:3],16),int(c[3:5],16),int(c[5:7],16)
        return "#{:02x}{:02x}{:02x}".format(
            max(0,int(r*.82)),max(0,int(g*.82)),max(0,int(b*.82)))

    def _draw(self, color=None):
        if color is None: color = self._bg if self._on else self._DISABLED
        self.delete('all')
        w, h = self.winfo_width(), self.winfo_height()
        if w < 4 or h < 4: return
        r = min(self._R, w//2, h//2)
        self.create_rectangle(r, 0, w-r, h, fill=color, outline='')
        self.create_rectangle(0, r, w, h-r, fill=color, outline='')
        for cx,cy,st in [(0,0,90),(w-2*r,0,0),(0,h-2*r,180),(w-2*r,h-2*r,270)]:
            self.create_arc(cx,cy,cx+2*r,cy+2*r,
                            start=st,extent=90,fill=color,outline=color)
        self.create_text(w//2, h//2, text=self._t,
                         font=self._fnt, fill='#ffffff', anchor='center')

    def _on_cfg(self, e=None):
        # winfo_width returns 1 until first layout — use event width if available
        w = (e.width if e and hasattr(e,'width') else 0) or self.winfo_width()
        if w > 4:
            self._draw()
        else:
            self.after(20, self._draw)   # retry after layout pass

    def _hover_in(self, _=None):
        if self._on: self._draw(self._dk(self._bg))
    def _hover_out(self, _=None): self._draw()
    def _click(self, _=None):
        if self._on and self._cmd:
            self._draw(self._dk(self._dk(self._bg)))
            self.after(120, self._draw)
            self._cmd()

    def enable(self):
        self._on = True
        self.config(cursor='hand2')
        self.after_idle(self._draw)   # draw after Tk processes geometry

    def disable(self):
        self._on = False
        self.config(cursor='arrow')
        self.after_idle(self._draw)


def _btn(parent, text, cmd, bg, icon="", font_size=10, btn_pady=9, **kw):
    """Rounded button — always enabled."""
    kw.pop('padx', None); kw.pop('pady', None)
    b = RoundedButton(parent, text, cmd, bg, icon=icon,
                      font_size=font_size, btn_pady=btn_pady, **kw)
    b.enable()
    return b






def _setup_styles():
    s=ttk.Style()
    try: s.theme_use('clam')
    except: pass
    s.configure("POS.Treeview",background="#fff",foreground="#111827",
        fieldbackground="#fff",rowheight=34,font=(FONT,10),borderwidth=0)
    s.configure("POS.Treeview.Heading",background="#111827",foreground="#fff",
        font=(FONT,9,'bold'),relief='flat',padding=(8,6))
    s.map("POS.Treeview",background=[('selected','#2563EB')],foreground=[('selected','#fff')])

def _center(win,w,h):
    win.update_idletasks()
    sw,sh=win.winfo_screenwidth(),win.winfo_screenheight()
    win.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")

class CreditsModule:
    def __init__(self,parent,db_manager):
        self.parent=parent; self.db=db_manager
        self.selected_credit_id=None
        _setup_styles()
        self._build(); self.load_credits()

    def _stat_card(self,parent,title,value,accent):
        outer=RoundedCard(parent,padx=16,pady=14)
        inner=outer.body
        tk.Label(inner,text=title,font=(FONT,9),bg=THEME["card_bg"],fg=THEME["txt_secondary"]).pack(anchor='w')
        lbl=tk.Label(inner,text=value,font=(FONT,18,'bold'),bg=THEME["card_bg"],fg=accent)
        lbl.pack(anchor='w',pady=(4,0))
        tk.Frame(inner,bg=accent,height=3).pack(fill='x',pady=(10,0))
        inner.value_label=lbl
        outer.value_label=lbl
        return outer

    def _build(self):
        bg=THEME["ct_bg"]
        # Header
        hdr=tk.Frame(self.parent,bg=bg)
        hdr.pack(fill='x',padx=28,pady=(20,0))
        tk.Label(hdr,text="💳  Gestión de Créditos",font=(FONT,16,'bold'),
                 bg=bg,fg=THEME["txt_primary"]).pack(side='left')
        tk.Frame(self.parent,bg=THEME["card_border"],height=1).pack(fill='x',padx=28,pady=(10,14))

        # Filtros
        flt=tk.Frame(self.parent,bg=bg,padx=28)
        flt.pack(fill='x',pady=(0,12))
        tk.Label(flt,text="Filtrar:",font=(FONT,10,'bold'),bg=bg,fg=THEME["txt_secondary"]).pack(side='left',padx=(0,10))
        self.status_var=tk.StringVar(value='all')
        for val,lbl in [('all','Todos'),('active','Activos'),('paid','Pagados'),('overdue','Vencidos')]:
            rb=tk.Radiobutton(flt,text=lbl,variable=self.status_var,value=val,
               font=(FONT,10),bg=bg,fg=THEME["txt_primary"],selectcolor=bg,
               command=self.load_credits)
            rb.pack(side='left',padx=(0,16))

        # Cards estadísticas
        stats_row=tk.Frame(self.parent,bg=bg)
        stats_row.pack(fill='x',padx=24,pady=(0,14))
        self.card_active  =self._stat_card(stats_row,"Créditos Activos","0",THEME["acc_blue"])
        self.card_active.grid(row=0,column=0,sticky='nsew',padx=4)
        self.card_pending =self._stat_card(stats_row,"Total por Cobrar","Gs. 0",THEME["acc_amber"])
        self.card_pending.grid(row=0,column=1,sticky='nsew',padx=4)
        self.card_overdue =self._stat_card(stats_row,"Vencidos","0",THEME["acc_rose"])
        self.card_overdue.grid(row=0,column=2,sticky='nsew',padx=4)
        for i in range(3): stats_row.columnconfigure(i,weight=1)

        # Tabla
        tbl_outer=RoundedCard(self.parent,padx=0,pady=0,fill_mode=True)
        tbl_outer.pack(fill='both',expand=True,padx=24,pady=(0,10))
        tbl=tbl_outer.body

        cols=('ID','Cliente','Productos','Total','Saldo','Cuota','Pagadas','Próximo','Estado')
        self.tree=ttk.Treeview(tbl,columns=cols,show='headings',style='POS.Treeview')
        for col,w in zip(cols,[40,140,220,100,100,90,60,100,75]):
            self.tree.heading(col,text=col)
            self.tree.column(col,width=w,anchor='center' if col not in('Cliente',) else 'w')
        sb=ttk.Scrollbar(tbl,orient='vertical',command=self.tree.yview)
        self.tree.configure(yscrollcommand=sb.set)
        sb.pack(side='right',fill='y',padx=(0,4),pady=4)
        self.tree.pack(fill='both',expand=True,padx=4,pady=4)
        self.tree.bind('<<TreeviewSelect>>',self._on_select)
        self.tree.bind('<Double-Button-1>',lambda _:self.show_credit_details())
        self.tree.tag_configure('active',background=THEME["row_even"])
        self.tree.tag_configure('paid',background='#F0FDF4')
        self.tree.tag_configure('overdue',background='#FFF1F2')

        # Botones acción
        act=tk.Frame(self.parent,bg=bg,padx=24,pady=10)
        act.pack(fill='x')
        self.btn_pay=_btn(act,"Registrar Pago",self.show_payment_dialog,THEME["acc_green"],"💰")
        self.btn_pay.pack(side='left',padx=(0,8))
        self.btn_hist=_btn(act,"Ver Historial",self.show_payment_history,THEME["acc_blue"],"📋")
        self.btn_hist.pack(side='left',padx=(0,8))
        self.btn_date=_btn(act,"Cambiar Fecha",self.show_change_date_dialog,THEME["btn_secondary"],"📅")
        self.btn_date.pack(side='left',padx=(0,8))
        btn_exp=_btn(act,"Exportar",self.show_export_dialog,THEME["acc_purple"],"📥")
        btn_exp.enable(); btn_exp.pack(side='right')

    def show_export_dialog(self):
        """Diálogo para exportar créditos a Excel o PDF."""
        dlg=tk.Toplevel(self.parent)
        _set_icon(dlg); dlg.title("Exportar Créditos")
        dlg.configure(bg=THEME["ct_bg"]); dlg.resizable(False,False)
        dlg.transient(self.parent); dlg.grab_set(); _center(dlg,460,420)

        btn_bar=tk.Frame(dlg,bg=THEME["card_bg"],padx=16,pady=12)
        btn_bar.pack(fill='x',side='bottom')
        tk.Frame(dlg,bg=THEME["card_border"],height=1).pack(fill='x',side='bottom')

        hdr=tk.Frame(dlg,bg=THEME["sb_bg"],pady=12); hdr.pack(fill='x')
        tk.Label(hdr,text="📥  Exportar Créditos",font=(FONT,13,'bold'),
                 bg=THEME["sb_bg"],fg="#fff").pack(anchor='w',padx=16)

        body=tk.Frame(dlg,bg=THEME["ct_bg"],padx=24,pady=16); body.pack(fill='both',expand=True)

        # ── Formato ──
        tk.Label(body,text="Formato de exportación:",font=(FONT,10,'bold'),
                 bg=THEME["ct_bg"],fg=THEME["txt_secondary"]).pack(anchor='w',pady=(0,6))
        fmt_var=tk.StringVar(value="excel")
        fmt_f=tk.Frame(body,bg=THEME["ct_bg"]); fmt_f.pack(fill='x',pady=(0,14))
        for val,lbl,ico in [("excel","Excel (.xlsx)","📊"),("pdf","PDF (.pdf)","📄")]:
            tk.Radiobutton(fmt_f,text=f"{ico}  {lbl}",variable=fmt_var,value=val,
               font=(FONT,10),bg=THEME["ct_bg"],selectcolor=THEME["ct_bg"],
               fg=THEME["txt_primary"]).pack(side='left',padx=(0,24))

        # ── Clientes ──
        tk.Label(body,text="¿Qué clientes exportar?",font=(FONT,10,'bold'),
                 bg=THEME["ct_bg"],fg=THEME["txt_secondary"]).pack(anchor='w',pady=(0,6))
        scope_var=tk.StringVar(value="all")
        tk.Radiobutton(body,text="📋  Todos los clientes",variable=scope_var,value="all",
           font=(FONT,10),bg=THEME["ct_bg"],selectcolor=THEME["ct_bg"],
           fg=THEME["txt_primary"]).pack(anchor='w',pady=2)

        sel_rb=tk.Radiobutton(body,text="👤  Solo cliente seleccionado",variable=scope_var,value="selected",
           font=(FONT,10),bg=THEME["ct_bg"],selectcolor=THEME["ct_bg"],
           fg=THEME["txt_primary"]); sel_rb.pack(anchor='w',pady=2)
        if not self.selected_credit_id: sel_rb.config(state='disabled')

        # ── Filtro estado ──
        tk.Label(body,text="Filtrar por estado:",font=(FONT,10,'bold'),
                 bg=THEME["ct_bg"],fg=THEME["txt_secondary"]).pack(anchor='w',pady=(12,6))
        est_var=tk.StringVar(value="all")
        ef=tk.Frame(body,bg=THEME["ct_bg"]); ef.pack(fill='x',pady=(0,8))
        for val,lbl in [("all","Todos"),("active","Activos"),("paid","Pagados")]:
            tk.Radiobutton(ef,text=lbl,variable=est_var,value=val,
               font=(FONT,10),bg=THEME["ct_bg"],selectcolor=THEME["ct_bg"]).pack(side='left',padx=(0,16))

        lbl_status=tk.Label(body,text="",font=(FONT,9),bg=THEME["ct_bg"],fg=THEME["acc_green"])
        lbl_status.pack(anchor='w',pady=(8,0))

        def do_export():
            fmt=fmt_var.get(); scope=scope_var.get(); est=est_var.get()
            try:
                credits=self.db.get_all_credit_sales() if est=='all'                     else self.db.get_all_credit_sales(est)
                if scope=="selected" and self.selected_credit_id:
                    credits=[c for c in credits if c.get('id')==self.selected_credit_id]
                if not credits:
                    lbl_status.config(text="⚠  No hay créditos para exportar.",fg=THEME["acc_rose"]); return
                # Obtener detalle de productos por crédito
                for c in credits:
                    vid=c.get('venta_id') or c.get('sale_id')
                    c['items']=self.db.get_sale_detail(int(vid)) if vid else []
                    c['payments']=self.db.get_credit_payments(c['id'])
                lbl_status.config(text="⏳  Generando archivo...",fg=THEME["acc_amber"])
                dlg.update()
                if fmt=="excel":
                    path=self._export_excel(credits)
                else:
                    path=self._export_pdf(credits)
                lbl_status.config(text=f"✅  Guardado en: {path}",fg=THEME["acc_green"])
                import os
                os.startfile(os.path.dirname(path))
            except Exception as ex:
                lbl_status.config(text=f"❌  Error: {ex}",fg=THEME["acc_rose"])

        b1=_btn(btn_bar,"Exportar",do_export,THEME["acc_purple"],"📥"); b1.enable(); b1.pack(side='left',padx=(0,8))
        b2=_btn(btn_bar,"Cancelar",dlg.destroy,THEME["btn_secondary"],"✕"); b2.enable(); b2.pack(side='left')

    def _export_path(self, ext):
        """Devuelve ruta de exportación en Documentos del usuario."""
        import os, sys, datetime
        docs = os.path.join(os.path.expanduser("~"), "Documents")
        os.makedirs(docs, exist_ok=True)
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        return os.path.join(docs, f"creditos_{ts}.{ext}")

    def _export_excel(self, credits):
        """Exporta créditos a .xlsx con hoja resumen y detalle de productos."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()

        # Datos empresa
        co = _get_company()
        empresa  = co.get("razon_social","Venialgo POS")
        ruc      = co.get("ruc","")
        telefono = co.get("telefono","")
        direccion= co.get("direccion","")
        correo   = co.get("correo","")

        # Colores
        HDR_BG  = "111827"; HDR_FG  = "FFFFFF"
        SUB_BG  = "1E3A5F"; SUB_FG  = "FFFFFF"
        ROW1_BG = "F9FAFB"; ROW2_BG = "FFFFFF"
        GRN_BG  = "D1FAE5"; RED_BG  = "FFF1F2"
        ACCENT  = "2563EB"
        thin = Side(style='thin', color='D1D5DB')
        brd  = Border(left=thin,right=thin,top=thin,bottom=thin)

        def hdr_style(cell, bg=HDR_BG, fg=HDR_FG):
            cell.font      = Font(name='Arial',bold=True,color=fg,size=10)
            cell.fill      = PatternFill('solid',start_color=bg)
            cell.alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
            cell.border    = brd

        def data_style(cell, bold=False, center=False, bg=None):
            cell.font      = Font(name='Arial',bold=bold,size=9)
            cell.alignment = Alignment(horizontal='center' if center else 'left',vertical='center')
            cell.border    = brd
            if bg: cell.fill = PatternFill('solid',start_color=bg)

        # ── HOJA 1: RESUMEN ──────────────────────────────
        ws1 = wb.active; ws1.title = "Resumen de Créditos"
        ws1.freeze_panes = "A5"

        from datetime import datetime as _dt
        # Fila 1: Nombre empresa
        ws1.merge_cells("A1:J1")
        ws1["A1"] = empresa
        ws1["A1"].font = Font(name='Arial',bold=True,size=14,color=ACCENT)
        ws1["A1"].alignment = Alignment(horizontal='center',vertical='center')
        ws1.row_dimensions[1].height = 24

        # Fila 2: RUC | Tel | Dirección | Correo
        info_parts = []
        if ruc:      info_parts.append(f"RUC: {ruc}")
        if telefono: info_parts.append(f"Tel: {telefono}")
        if correo:   info_parts.append(f"Email: {correo}")
        if direccion:info_parts.append(direccion)
        ws1.merge_cells("A2:J2")
        ws1["A2"] = "   |   ".join(info_parts)
        ws1["A2"].font = Font(name='Arial',size=9,color="374151")
        ws1["A2"].alignment = Alignment(horizontal='center')
        ws1.row_dimensions[2].height = 16

        # Fila 3: Título reporte
        ws1.merge_cells("A3:J3")
        ws1["A3"] = "REPORTE DE CRÉDITOS"
        ws1["A3"].font = Font(name='Arial',bold=True,size=12,color="111827")
        ws1["A3"].alignment = Alignment(horizontal='center',vertical='center')
        ws1["A3"].fill = PatternFill('solid',start_color="E5E7EB")
        ws1.row_dimensions[3].height = 20

        # Fila 4: Fecha y total
        ws1.merge_cells("A4:J4")
        ws1["A4"] = f"Generado: {_dt.now().strftime('%d/%m/%Y %H:%M')}   |   Total créditos: {len(credits)}"
        ws1["A4"].font = Font(name='Arial',italic=True,size=9,color="6B7280")
        ws1["A4"].alignment = Alignment(horizontal='center')
        ws1.row_dimensions[4].height = 14

        cols_res = ["ID","Cliente","Teléfono","Total Venta","Cuota Inicial",
                    "Saldo Restante","Cuota Mensual","Cuotas Pagadas","Próximo Pago","Estado"]
        for ci,col in enumerate(cols_res,1):
            c = ws1.cell(row=5,column=ci,value=col); hdr_style(c)

        widths_res = [7,22,16,15,15,15,14,14,14,12]
        for ci,w in enumerate(widths_res,1):
            ws1.column_dimensions[get_column_letter(ci)].width = w
        ws1.row_dimensions[5].height = 22

        for ri,cr in enumerate(credits,6):
            bg = GRN_BG if cr.get('status')=='paid' else (RED_BG if cr.get('status')=='overdue' else ROW1_BG if ri%2==0 else ROW2_BG)
            vals = [cr.get('id',''), cr.get('customer_name',''), cr.get('phone',''),
                    cr.get('total_amount',0), cr.get('down_payment',0),
                    cr.get('remaining_balance',0), cr.get('installment_amount',0),
                    f"{cr.get('paid_installments',0)}/{cr.get('total_installments',0)}",
                    cr.get('next_payment_date','—'),
                    'Pagado' if cr.get('status')=='paid' else 'Activo']
            for ci,v in enumerate(vals,1):
                cell = ws1.cell(row=ri,column=ci,value=v)
                data_style(cell, center=ci not in(2,3), bg=bg[2:] if bg.startswith('#') else bg)
            ws1.row_dimensions[ri].height = 18

        # Fila totales
        last = 5+len(credits)
        ws1.cell(row=last+1,column=1,value="TOTAL").font = Font(bold=True,name='Arial',size=10)
        for ci,col in [(4,'total_amount'),(5,'down_payment'),(6,'remaining_balance'),(7,'installment_amount')]:
            total=sum(c.get(col,0) for c in credits)
            cell=ws1.cell(row=last+1,column=ci,value=total)
            data_style(cell,bold=True,center=True)

        # ── HOJA 2: DETALLE DE PRODUCTOS ─────────────────
        ws2 = wb.create_sheet("Detalle de Productos")
        ws2.freeze_panes = "A4"
        ws2.merge_cells("A1:G1")
        ws2["A1"] = empresa
        ws2["A1"].font = Font(name='Arial',bold=True,size=13,color=ACCENT)
        ws2["A1"].alignment = Alignment(horizontal='center',vertical='center')
        ws2.row_dimensions[1].height = 22
        ws2.merge_cells("A2:G2")
        ws2["A2"] = "DETALLE DE PRODUCTOS POR CRÉDITO"
        ws2["A2"].font = Font(name='Arial',bold=True,size=11,color="111827")
        ws2["A2"].fill = PatternFill('solid',start_color="E5E7EB")
        ws2["A2"].alignment = Alignment(horizontal='center',vertical='center')
        ws2.row_dimensions[2].height = 18

        cols_det = ["ID Crédito","Cliente","Producto","Cantidad","Precio Unit.","Subtotal","Estado Crédito"]
        for ci,col in enumerate(cols_det,1):
            c=ws2.cell(row=3,column=ci,value=col); hdr_style(c)
        for ci,w in enumerate([12,22,28,10,14,14,14],1):
            ws2.column_dimensions[get_column_letter(ci)].width = w
        ws2.row_dimensions[3].height = 22

        ri2=4
        for cr in credits:
            items=cr.get('items',[])
            estado='Pagado' if cr.get('status')=='paid' else 'Activo'
            bg2 = GRN_BG if cr.get('status')=='paid' else ROW1_BG
            if not items:
                row_vals=[cr.get('id',''),cr.get('customer_name',''),'Sin detalle','—','—','—',estado]
                for ci,v in enumerate(row_vals,1):
                    cell=ws2.cell(row=ri2,column=ci,value=v)
                    data_style(cell,center=ci not in(2,3),bg=bg2)
                ri2+=1
            else:
                for it in items:
                    nombre=it.get('nombre') or it.get('name','—')
                    cant  =it.get('cantidad') or it.get('quantity',1)
                    precio=it.get('precio') or it.get('unit_price',0)
                    sub   =it.get('subtotal',0)
                    row_vals=[cr.get('id',''),cr.get('customer_name',''),nombre,cant,precio,sub,estado]
                    for ci,v in enumerate(row_vals,1):
                        cell=ws2.cell(row=ri2,column=ci,value=v)
                        data_style(cell,center=ci not in(2,3),bg=bg2)
                    ri2+=1
            ws2.row_dimensions[ri2-1].height=18

        # ── HOJA 3: PAGOS REALIZADOS ─────────────────────
        ws3=wb.create_sheet("Historial de Pagos")
        ws3.freeze_panes="A4"
        ws3.merge_cells("A1:F1")
        ws3["A1"]=empresa
        ws3["A1"].font=Font(name='Arial',bold=True,size=13,color=ACCENT)
        ws3["A1"].alignment=Alignment(horizontal='center',vertical='center')
        ws3.row_dimensions[1].height=22
        ws3.merge_cells("A2:F2")
        ws3["A2"]="HISTORIAL DE PAGOS"
        ws3["A2"].font=Font(name='Arial',bold=True,size=11,color="111827")
        ws3["A2"].fill=PatternFill('solid',start_color="E5E7EB")
        ws3["A2"].alignment=Alignment(horizontal='center',vertical='center')
        ws3.row_dimensions[2].height=18
        cols_pag=["ID Crédito","Cliente","Fecha Pago","Monto","Nota","Estado Crédito"]
        for ci,col in enumerate(cols_pag,1):
            c=ws3.cell(row=3,column=ci,value=col); hdr_style(c)
        for ci,w in enumerate([12,22,16,14,30,14],1):
            ws3.column_dimensions[get_column_letter(ci)].width=w
        ws3.row_dimensions[3].height=22
        ri3=4
        for cr in credits:
            estado='Pagado' if cr.get('status')=='paid' else 'Activo'
            pagos=cr.get('payments',[])
            if not pagos:
                row_vals=[cr.get('id',''),cr.get('customer_name',''),'Sin pagos','—','—',estado]
                for ci,v in enumerate(row_vals,1):
                    cell=ws3.cell(row=ri3,column=ci,value=v); data_style(cell,center=ci not in(2,5))
                ri3+=1
            else:
                for pg in pagos:
                    row_vals=[cr.get('id',''),cr.get('customer_name',''),
                              pg.get('payment_date',''),pg.get('amount',0),
                              pg.get('notes','') or '—',estado]
                    for ci,v in enumerate(row_vals,1):
                        cell=ws3.cell(row=ri3,column=ci,value=v); data_style(cell,center=ci not in(2,5))
                    ri3+=1

        path=self._export_path("xlsx")
        wb.save(path)
        return path

    def _export_pdf(self, credits):
        """Exporta créditos a PDF usando reportlab."""
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib import colors
            from reportlab.lib.units import cm
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.enums import TA_CENTER, TA_LEFT
            from datetime import datetime as _dt

            path=self._export_path("pdf")
            doc=SimpleDocTemplate(path,pagesize=landscape(A4),
                leftMargin=1.5*cm,rightMargin=1.5*cm,topMargin=2*cm,bottomMargin=2*cm)
            styles=getSampleStyleSheet()
            DARK=colors.HexColor("#111827"); BLUE=colors.HexColor("#2563EB")
            GRN =colors.HexColor("#D1FAE5"); RED =colors.HexColor("#FFF1F2")
            HDRC=colors.HexColor("#111827")

            title_style=ParagraphStyle('title',fontSize=16,textColor=BLUE,
                                        spaceAfter=4,alignment=TA_CENTER,fontName='Helvetica-Bold')
            sub_style  =ParagraphStyle('sub',  fontSize=9, textColor=colors.grey,
                                        spaceAfter=12,alignment=TA_CENTER)
            sec_style  =ParagraphStyle('sec',  fontSize=11,textColor=DARK,
                                        spaceBefore=14,spaceAfter=6,fontName='Helvetica-Bold')

            co2 = _get_company()
            emp2     = co2.get("razon_social","Venialgo POS")
            ruc2     = co2.get("ruc","")
            tel2     = co2.get("telefono","")
            dir2     = co2.get("direccion","")
            mail2    = co2.get("correo","")
            info2_parts=[]
            if ruc2:  info2_parts.append(f"RUC: {ruc2}")
            if tel2:  info2_parts.append(f"Tel: {tel2}")
            if mail2: info2_parts.append(f"Email: {mail2}")
            if dir2:  info2_parts.append(dir2)

            story=[]
            story.append(Paragraph(emp2, title_style))
            if info2_parts:
                story.append(Paragraph("  |  ".join(info2_parts), sub_style))
            story.append(Paragraph("REPORTE DE CRÉDITOS", ParagraphStyle('rep',fontSize=12,
                textColor=DARK,spaceAfter=8,alignment=TA_CENTER,fontName='Helvetica-Bold')))
            story.append(Paragraph(
                f"Generado: {_dt.now().strftime('%d/%m/%Y %H:%M')}  |  Total créditos: {len(credits)}", sub_style))

            # ── TABLA RESUMEN ──
            story.append(Paragraph("Resumen de Créditos", sec_style))
            hdr_res=["ID","Cliente","Total Venta","Cuota Inicial","Saldo","Cuota","Cuotas","Próx. Pago","Estado"]
            rows_res=[hdr_res]
            for cr in credits:
                rows_res.append([
                    str(cr.get('id','')), cr.get('customer_name','')[:28],
                    format_currency(cr.get('total_amount',0)),
                    format_currency(cr.get('down_payment',0)),
                    format_currency(cr.get('remaining_balance',0)),
                    format_currency(cr.get('installment_amount',0)),
                    f"{cr.get('paid_installments',0)}/{cr.get('total_installments',0)}",
                    cr.get('next_payment_date','—'),
                    'Pagado' if cr.get('status')=='paid' else 'Activo'
                ])
            col_w=[1.2*cm,5*cm,3.2*cm,3*cm,3*cm,3*cm,2.2*cm,3*cm,2.2*cm]
            t1=Table(rows_res,colWidths=col_w,repeatRows=1)
            ts1=[
                ('BACKGROUND',(0,0),(-1,0),HDRC),('TEXTCOLOR',(0,0),(-1,0),colors.white),
                ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,0),8),
                ('FONTNAME',(0,1),(-1,-1),'Helvetica'),('FONTSIZE',(0,1),(-1,-1),8),
                ('ALIGN',(0,0),(-1,-1),'CENTER'),('ALIGN',(1,1),(1,-1),'LEFT'),
                ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor("#F9FAFB")]),
                ('GRID',(0,0),(-1,-1),0.4,colors.HexColor("#D1D5DB")),
                ('TOPPADDING',(0,0),(-1,-1),5),('BOTTOMPADDING',(0,0),(-1,-1),5),
            ]
            for ri,cr in enumerate(credits,1):
                if cr.get('status')=='paid':
                    ts1.append(('BACKGROUND',(0,ri),(-1,ri),GRN))
            t1.setStyle(TableStyle(ts1)); story.append(t1)

            # ── TABLA PRODUCTOS ──
            story.append(Paragraph("Detalle de Productos", sec_style))
            hdr_prod=["ID Cred.","Cliente","Producto","Cant.","Precio Unit.","Subtotal"]
            rows_prod=[hdr_prod]
            for cr in credits:
                items=cr.get('items',[])
                if not items:
                    rows_prod.append([str(cr.get('id','')),cr.get('customer_name','')[:28],'Sin detalle','—','—','—'])
                else:
                    for it in items:
                        rows_prod.append([
                            str(cr.get('id','')), cr.get('customer_name','')[:28],
                            (it.get('nombre') or it.get('name','—'))[:32],
                            str(it.get('cantidad') or it.get('quantity',1)),
                            format_currency(it.get('precio') or it.get('unit_price',0)),
                            format_currency(it.get('subtotal',0))
                        ])
            col_w2=[1.5*cm,5.5*cm,7*cm,2*cm,3.5*cm,3.5*cm]
            t2=Table(rows_prod,colWidths=col_w2,repeatRows=1)
            t2.setStyle(TableStyle([
                ('BACKGROUND',(0,0),(-1,0),HDRC),('TEXTCOLOR',(0,0),(-1,0),colors.white),
                ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,0),8),
                ('FONTNAME',(0,1),(-1,-1),'Helvetica'),('FONTSIZE',(0,1),(-1,-1),8),
                ('ALIGN',(0,0),(-1,-1),'CENTER'),('ALIGN',(2,1),(2,-1),'LEFT'),('ALIGN',(1,1),(1,-1),'LEFT'),
                ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor("#F9FAFB")]),
                ('GRID',(0,0),(-1,-1),0.4,colors.HexColor("#D1D5DB")),
                ('TOPPADDING',(0,0),(-1,-1),5),('BOTTOMPADDING',(0,0),(-1,-1),5),
            ])); story.append(t2)

            doc.build(story)
            return path
        except ImportError:
            # Fallback si reportlab no está disponible
            return self._export_pdf_fallback(credits)

    def _export_pdf_fallback(self, credits):
        """Fallback: genera HTML que el usuario puede imprimir como PDF."""
        from datetime import datetime as _dt
        path=self._export_path("html")
        rows=""
        for cr in credits:
            est='Pagado' if cr.get('status')=='paid' else 'Activo'
            color='#D1FAE5' if cr.get('status')=='paid' else '#FFFFFF'
            items=cr.get('items',[])
            prods=', '.join((it.get('nombre') or it.get('name',''))[:20] for it in items) or '—'
            rows+=f"""<tr style="background:{color}">
                <td>{cr.get('id','')}</td><td>{cr.get('customer_name','')}</td>
                <td>{prods}</td>
                <td>{format_currency(cr.get('total_amount',0))}</td>
                <td>{format_currency(cr.get('remaining_balance',0))}</td>
                <td>{cr.get('next_payment_date','—')}</td><td>{est}</td></tr>"""
        co3=_get_company()
        emp3=co3.get("razon_social","Venialgo POS")
        info3=[]
        if co3.get("ruc"):     info3.append(f"RUC: {co3['ruc']}")
        if co3.get("telefono"):info3.append(f"Tel: {co3['telefono']}")
        if co3.get("correo"):  info3.append(f"Email: {co3['correo']}")
        if co3.get("direccion"):info3.append(co3['direccion'])
        info3_str=" &nbsp;|&nbsp; ".join(info3)
        html=f"""<!DOCTYPE html><html><head><meta charset="utf-8">
        <title>Créditos</title>
        <style>body{{font-family:Arial;font-size:11px;margin:20px}}
        h1{{color:#2563EB;font-size:18px;margin-bottom:4px}}
        .info{{color:#374151;font-size:10px;margin-bottom:12px}}
        h2{{color:#111827;font-size:13px;border-bottom:2px solid #2563EB;padding-bottom:4px}}
        table{{border-collapse:collapse;width:100%}}
        th{{background:#111827;color:white;padding:6px 8px;text-align:center}}
        td{{padding:5px 8px;border:1px solid #D1D5DB;text-align:center}}
        </style></head><body>
        <h1>{emp3}</h1>
        <p class="info">{info3_str}</p>
        <h2>Reporte de Créditos</h2>
        <p>Generado: {_dt.now().strftime('%d/%m/%Y %H:%M')} | Total: {len(credits)} créditos</p>
        <table><tr><th>ID</th><th>Cliente</th><th>Productos</th>
        <th>Total</th><th>Saldo</th><th>Próx.Pago</th><th>Estado</th></tr>
        {rows}</table>
        <p style="margin-top:20px;color:gray">Para exportar a PDF: Archivo > Imprimir > Guardar como PDF</p>
        </body></html>"""
        with open(path,'w',encoding='utf-8') as f: f.write(html)
        import os; os.startfile(path)
        return path

    def load_credits(self):
        for i in self.tree.get_children(): self.tree.delete(i)
        sf=self.status_var.get()
        credits=self.db.get_all_credit_sales() if sf=='all' \
            else self.db.get_overdue_credits() if sf=='overdue' \
            else self.db.get_all_credit_sales(sf)
        active_count=0; total_pending=0; overdue_count=0
        today=datetime.now().strftime('%Y-%m-%d')
        for c in credits:
            status=c.get('status','active')
            nxt=c.get('next_payment_date','')
            is_overdue=status=='active' and nxt and nxt<today
            if is_overdue: overdue_count+=1; tag='overdue'
            elif status=='active': tag='active'
            else: tag='paid'
            if status=='active': active_count+=1; total_pending+=c.get('remaining_balance',0)
            status_txt='Vencido' if is_overdue else ('Activo' if status=='active' else 'Pagado')
            # Obtener productos del crédito
            prods_str = '—'
            try:
                vid = c.get('venta_id') or c.get('sale_id')
                if vid:
                    items = self.db.get_sale_detail(int(vid))
                    if items:
                        nombres = [it.get('nombre') or it.get('name','') for it in items]
                        prods_str = ', '.join(n for n in nombres if n)[:45]
                        if not prods_str: prods_str = '—'
            except Exception:
                pass
            pagadas_txt = f"{c.get('paid_installments',0)}/{c.get('total_installments',0)}"
            self.tree.insert('','end',tags=(tag,),values=(
                c.get('id',''), c.get('customer_name',''), prods_str,
                format_currency(c.get('total_amount',0)),
                format_currency(c.get('remaining_balance',0)),
                format_currency(c.get('installment_amount',0)),
                pagadas_txt, nxt or '—', status_txt))
        self.card_active.value_label.config(text=str(active_count))
        self.card_pending.value_label.config(text=format_currency(total_pending))
        self.card_overdue.value_label.config(text=str(overdue_count))

    def _on_select(self,_=None):
        sel=self.tree.selection()
        if sel:
            self.selected_credit_id=self.tree.item(sel[0])['values'][0]
            status=self.tree.item(sel[0])['values'][8]
            can=status in('Activo','Vencido')
            self.btn_pay.enable() if can else self.btn_pay.disable()
            self.btn_date.enable() if can else self.btn_date.disable()
            self.btn_hist.enable()
        else:
            self.selected_credit_id=None
            for b in [self.btn_pay,self.btn_hist,self.btn_date]: b.disable()

    def show_credit_details(self):
        if not self.selected_credit_id: return
        c=self.db.get_credit_sale_by_id(self.selected_credit_id)
        if not c: return
        dlg=tk.Toplevel(self.parent)
        _set_icon(dlg); dlg.title(f"Crédito #{c['id']} — Detalle")
        dlg.configure(bg=THEME["ct_bg"]); dlg.resizable(True,True)
        dlg.transient(self.parent); _center(dlg,580,600)
        dlg.minsize(500,500)

        # Header
        hdr=tk.Frame(dlg,bg=THEME["sb_bg"],pady=14)
        hdr.pack(fill='x')
        tk.Label(hdr,text=f"💳  Crédito #{c['id']} — {c.get('customer_name','')}",
                 font=(FONT,13,'bold'),bg=THEME["sb_bg"],fg="#fff").pack(anchor='w',padx=20)
        estado_txt='Activo' if c.get('status')=='active' else 'Pagado'
        estado_col=THEME["acc_green"] if c.get('status')=='active' else THEME["acc_blue"]
        tk.Label(hdr,text=estado_txt,font=(FONT,9,'bold'),bg=THEME["sb_bg"],fg=estado_col).pack(anchor='w',padx=22)

        # Canvas scroll
        canvas=tk.Canvas(dlg,bg=THEME["ct_bg"],highlightthickness=0)
        sb_v=tk.Scrollbar(dlg,orient='vertical',command=canvas.yview)
        canvas.configure(yscrollcommand=sb_v.set)
        sb_v.pack(side='right',fill='y')
        canvas.pack(fill='both',expand=True)
        body=tk.Frame(canvas,bg=THEME["ct_bg"],padx=24,pady=16)
        canvas.create_window((0,0),window=body,anchor='nw',tags='body')
        body.bind('<Configure>',lambda e:canvas.configure(scrollregion=canvas.bbox('all')))
        canvas.bind('<Configure>',lambda e:canvas.itemconfig('body',width=e.width))

        # ── Resumen financiero ──
        sec=tk.Frame(body,bg=THEME["card_border"]); sec.pack(fill='x',pady=(0,12))
        inn=tk.Frame(sec,bg=THEME["card_bg"],padx=16,pady=12); inn.pack(fill='x',padx=1,pady=1)
        tk.Label(inn,text="Resumen Financiero",font=(FONT,10,'bold'),
                 bg=THEME["card_bg"],fg=THEME["acc_blue"]).pack(anchor='w',pady=(0,8))
        rows=[("Cliente:",          c.get('customer_name','')),
              ("Total de venta:",   format_currency(c.get('total_amount',0))),
              ("Cuota inicial:",    format_currency(c.get('down_payment',0))),
              ("Saldo restante:",   format_currency(c.get('remaining_balance',0))),
              ("Monto por cuota:",  format_currency(c.get('installment_amount',0))),
              ("Cuotas pagadas:",   f"{c.get('paid_installments',0)} de {c.get('total_installments',0)}"),
              ("Próximo pago:",     c.get('next_payment_date','—'))]
        grid=tk.Frame(inn,bg=THEME["card_bg"]); grid.pack(fill='x')
        for i,(lbl,val) in enumerate(rows):
            tk.Label(grid,text=lbl,font=(FONT,9,'bold'),bg=THEME["card_bg"],
                     fg=THEME["txt_secondary"],anchor='w',width=18).grid(row=i,column=0,sticky='w',pady=3)
            tk.Label(grid,text=str(val),font=(FONT,10),bg=THEME["card_bg"],
                     fg=THEME["txt_primary"],anchor='w').grid(row=i,column=1,sticky='w',padx=12,pady=3)

        # ── Productos adquiridos ──
        sec2=tk.Frame(body,bg=THEME["card_border"]); sec2.pack(fill='x',pady=(0,12))
        inn2=tk.Frame(sec2,bg=THEME["card_bg"],padx=16,pady=12); inn2.pack(fill='x',padx=1,pady=1)
        tk.Label(inn2,text="Productos Adquiridos",font=(FONT,10,'bold'),
                 bg=THEME["card_bg"],fg=THEME["acc_blue"]).pack(anchor='w',pady=(0,8))
        items=[]
        try:
            venta_id=c.get('venta_id') or c.get('sale_id')
            if venta_id:
                items=self.db.get_sale_detail(int(venta_id))
        except Exception: pass
        if items:
            prod_cols=('Producto','Cant.','Precio Unit.','Subtotal')
            ptree=ttk.Treeview(inn2,columns=prod_cols,show='headings',
                               style='POS.Treeview',height=min(len(items),6))
            for col,w in zip(prod_cols,[220,60,110,110]):
                ptree.heading(col,text=col)
                ptree.column(col,width=w,anchor='w' if col=='Producto' else 'center')
            for it in items:
                nombre=it.get('nombre') or it.get('name','—')
                cant  =it.get('cantidad') or it.get('quantity',1)
                precio=it.get('precio') or it.get('unit_price',0)
                sub   =it.get('subtotal',0)
                ptree.insert('','end',values=(nombre,cant,
                             format_currency(precio),format_currency(sub)))
            ptree.pack(fill='x')
        else:
            tk.Label(inn2,text="Sin detalle de productos registrado.",
                     font=(FONT,9),bg=THEME["card_bg"],fg=THEME["txt_secondary"]).pack(anchor='w')

        # ── Historial de pagos resumido ──
        try:
            pagos=self.db.get_credit_payments(c['id'])
            if pagos:
                sec3=tk.Frame(body,bg=THEME["card_border"]); sec3.pack(fill='x',pady=(0,12))
                inn3=tk.Frame(sec3,bg=THEME["card_bg"],padx=16,pady=12); inn3.pack(fill='x',padx=1,pady=1)
                tk.Label(inn3,text=f"Historial de Pagos ({len(pagos)} registros)",
                         font=(FONT,10,'bold'),bg=THEME["card_bg"],fg=THEME["acc_blue"]).pack(anchor='w',pady=(0,8))
                p_cols=('Fecha','Monto','Nota')
                pht=ttk.Treeview(inn3,columns=p_cols,show='headings',
                                 style='POS.Treeview',height=min(len(pagos),4))
                for col,w in zip(p_cols,[130,110,230]):
                    pht.heading(col,text=col); pht.column(col,width=w,anchor='center')
                for p in pagos[-8:]:
                    pht.insert('','end',values=(
                        p.get('payment_date',p.get('fecha','')),
                        format_currency(p.get('amount',p.get('monto',0))),
                        p.get('notes',p.get('nota','')) or '—'))
                pht.pack(fill='x')
        except Exception: pass

        # Botón cerrar
        bf=tk.Frame(dlg,bg=THEME["card_bg"],pady=10); bf.pack(fill='x')
        tk.Frame(dlg,bg=THEME["card_border"],height=1).pack(fill='x',side='bottom' if False else 'top')
        bc=_btn(bf,"Cerrar",dlg.destroy,THEME["btn_secondary"],"✕"); bc.enable(); bc.pack(pady=4)

    def show_payment_dialog(self):
        if not self.selected_credit_id: return
        c=self.db.get_credit_sale_by_id(self.selected_credit_id)
        if not c: return
        dlg=tk.Toplevel(self.parent)
        _set_icon(dlg); dlg.title("Registrar Pago")
        dlg.configure(bg=THEME["ct_bg"]); dlg.resizable(False,False)
        dlg.transient(self.parent); dlg.grab_set(); _center(dlg,420,440)

        # Botones PRIMERO
        btn_bar=tk.Frame(dlg,bg=THEME["card_bg"],padx=16,pady=12)
        btn_bar.pack(fill='x',side='bottom')
        tk.Frame(dlg,bg=THEME["card_border"],height=1).pack(fill='x',side='bottom')

        hdr=tk.Frame(dlg,bg=THEME["sb_bg"],pady=12)
        hdr.pack(fill='x')
        tk.Label(hdr,text=f"💰  Registrar Pago — {c.get('customer_name','')}",
                 font=(FONT,12,'bold'),bg=THEME["sb_bg"],fg="#fff").pack(anchor='w',padx=16)

        canvas=tk.Canvas(dlg,bg=THEME["ct_bg"],highlightthickness=0)
        canvas.pack(fill='both',expand=True)
        body=tk.Frame(canvas,bg=THEME["ct_bg"],padx=20,pady=16)
        canvas.create_window((0,0),window=body,anchor='nw',tags='body')
        def _rsz(e): canvas.itemconfig('body',width=e.width)
        def _srg(e): canvas.configure(scrollregion=canvas.bbox('all'))
        canvas.bind('<Configure>',_rsz); body.bind('<Configure>',_srg)
        tk.Label(body,text=f"Saldo actual: {format_currency(c.get('remaining_balance',0))}",
                 font=(FONT,12,'bold'),bg=THEME["ct_bg"],fg=THEME["acc_rose"]).pack(anchor='w',pady=(0,4))
        tk.Label(body,text=f"Cuota sugerida: {format_currency(c.get('installment_amount',0))}",
                 font=(FONT,10),bg=THEME["ct_bg"],fg=THEME["acc_green"]).pack(anchor='w',pady=(0,14))
        tk.Label(body,text="Monto a Pagar",font=(FONT,10,'bold'),
                 bg=THEME["ct_bg"],fg=THEME["txt_secondary"]).pack(anchor='w',pady=(0,4))
        amt_outer=tk.Frame(body,bg=THEME["input_brd"])
        amt_inner=tk.Frame(amt_outer,bg=THEME["input_bg"])
        amt_inner.pack(fill='x',padx=1,pady=1)
        amt_e=tk.Entry(amt_inner,font=(FONT,13),bg=THEME["input_bg"],
                       fg=THEME["txt_primary"],relief='flat',bd=0)
        amt_e.pack(fill='x',padx=12,pady=10)
        amt_outer.pack(fill='x',pady=(0,14))
        amt_e.insert(0,str(int(c.get('installment_amount',0)))); amt_e.focus()
        tk.Label(body,text="Notas (opcional)",font=(FONT,10,'bold'),
                 bg=THEME["ct_bg"],fg=THEME["txt_secondary"]).pack(anchor='w',pady=(0,4))
        notes_t=tk.Text(body,font=(FONT,10),height=3,
                        bg=THEME["input_bg"],relief='solid',bd=1,fg=THEME["txt_primary"])
        notes_t.pack(fill='x')

        def register():
            try:
                amt=float(amt_e.get())
                if amt<=0: messagebox.showerror("Error","Monto debe ser mayor a 0",parent=dlg); return
                if amt>c.get('remaining_balance',0):
                    if not messagebox.askyesno("Confirmar",f"Monto mayor al saldo. ¿Continuar?",parent=dlg): return
                self.db.add_credit_payment(self.selected_credit_id,amt,notes_t.get('1.0','end').strip())
                messagebox.showinfo("✅ Pago registrado","Pago registrado correctamente",parent=dlg)
                self.load_credits(); dlg.destroy()
            except ValueError: messagebox.showerror("Error","Ingrese un monto válido",parent=dlg)
            except Exception as e: messagebox.showerror("Error",str(e),parent=dlg)

        b1=_btn(btn_bar,"Registrar Pago",register,THEME["acc_green"],"💰"); b1.enable(); b1.pack(side='left',padx=(0,8))
        b2=_btn(btn_bar,"Cancelar",dlg.destroy,THEME["btn_secondary"],"✕"); b2.enable(); b2.pack(side='left')

    def show_payment_history(self):
        if not self.selected_credit_id: return
        c=self.db.get_credit_sale_by_id(self.selected_credit_id)
        payments=self.db.get_credit_payments(self.selected_credit_id)
        dlg=tk.Toplevel(self.parent)
        _set_icon(dlg); dlg.title("Historial de Pagos")
        dlg.configure(bg=THEME["ct_bg"]); dlg.resizable(True,True)
        dlg.transient(self.parent); dlg.grab_set(); _center(dlg,620,500)
        dlg.minsize(500,400)

        # Botón PRIMERO (regla tkinter side=bottom)
        btn_bar=tk.Frame(dlg,bg=THEME["card_bg"],padx=16,pady=10)
        btn_bar.pack(fill='x',side='bottom')
        tk.Frame(dlg,bg=THEME["card_border"],height=1).pack(fill='x',side='bottom')

        # Resumen totales SEGUNDO (side=bottom)
        sum_f=tk.Frame(dlg,bg=THEME["card_bg"],padx=16,pady=10)
        sum_f.pack(fill='x',side='bottom')
        tk.Frame(dlg,bg=THEME["card_border"],height=1).pack(fill='x',side='bottom')

        # Header
        hdr=tk.Frame(dlg,bg=THEME["sb_bg"],pady=12); hdr.pack(fill='x')
        tk.Label(hdr,text=f"📋  Historial — {c.get('customer_name','')}",
                 font=(FONT,12,'bold'),bg=THEME["sb_bg"],fg="#fff").pack(anchor='w',padx=16)
        tk.Label(hdr,text=f"{len(payments)} pago(s) registrado(s)",
                 font=(FONT,9),bg=THEME["sb_bg"],fg="#8AABD4").pack(anchor='w',padx=18)

        # Tabla
        tbl=tk.Frame(dlg,bg=THEME["card_bg"]); tbl.pack(fill='both',expand=True,padx=16,pady=12)
        cols=('Fecha','Monto','Notas')
        tree=ttk.Treeview(tbl,columns=cols,show='headings',style='POS.Treeview')
        for col,w in zip(cols,[170,130,280]):
            tree.heading(col,text=col); tree.column(col,width=w,anchor='w' if col=='Notas' else 'center')
        sb=ttk.Scrollbar(tbl,orient='vertical',command=tree.yview)
        tree.configure(yscrollcommand=sb.set)
        sb.pack(side='right',fill='y'); tree.pack(fill='both',expand=True)

        total_paid=0
        for i,p in enumerate(payments):
            tree.insert('','end',tags=('odd' if i%2 else 'even',),values=(
                format_date(p.get('payment_date','')),
                format_currency(p.get('amount',0)),
                p.get('notes','') or '—'))
            total_paid+=p.get('amount',0)
        tree.tag_configure('odd',background=THEME["row_odd"])
        tree.tag_configure('even',background=THEME["row_even"])

        # Rellenar totales
        tk.Label(sum_f,text=f"Total pagado: {format_currency(total_paid)}",
                 font=(FONT,10,'bold'),bg=THEME["card_bg"],fg=THEME["acc_green"]).pack(side='left')
        tk.Label(sum_f,text=f"Saldo restante: {format_currency(c.get('remaining_balance',0))}",
                 font=(FONT,10,'bold'),bg=THEME["card_bg"],fg=THEME["acc_rose"]).pack(side='right')

        # Botón cerrar habilitado
        bc=_btn(btn_bar,"Cerrar",dlg.destroy,THEME["btn_secondary"],"✕")
        bc.enable(); bc.pack(pady=2)
        dlg.bind('<Escape>', lambda e: dlg.destroy())

    def show_change_date_dialog(self):
        if not self.selected_credit_id: return
        c=self.db.get_credit_sale_by_id(self.selected_credit_id)
        dlg=tk.Toplevel(self.parent)
        _set_icon(dlg); dlg.title("Cambiar Fecha de Pago")
        dlg.configure(bg=THEME["ct_bg"]); dlg.resizable(False,False)
        dlg.transient(self.parent); dlg.grab_set(); _center(dlg,400,300)
        # Botones PRIMERO (side=bottom)
        btn_bar=tk.Frame(dlg,bg=THEME["card_bg"],padx=16,pady=12)
        btn_bar.pack(fill='x',side='bottom')
        tk.Frame(dlg,bg=THEME["card_border"],height=1).pack(fill='x',side='bottom')
        hdr=tk.Frame(dlg,bg=THEME["sb_bg"],pady=12); hdr.pack(fill='x')
        tk.Label(hdr,text=f"📅  Cambiar Fecha — {c.get('customer_name','')}",
                 font=(FONT,12,'bold'),bg=THEME["sb_bg"],fg="#fff").pack(anchor='w',padx=16)
        body=tk.Frame(dlg,bg=THEME["ct_bg"],padx=20,pady=16); body.pack(fill='both',expand=True)

        # Mostrar fecha actual en formato DD/MM/AAAA
        fecha_actual_iso = c.get('next_payment_date','')
        fecha_actual_display = ''
        try:
            from datetime import datetime as _dt
            fecha_actual_display = _dt.strptime(fecha_actual_iso,'%Y-%m-%d').strftime('%d/%m/%Y')
        except Exception:
            fecha_actual_display = fecha_actual_iso or '—'

        tk.Label(body,text=f"Fecha actual: {fecha_actual_display}",font=(FONT,10),
                 bg=THEME["ct_bg"],fg=THEME["txt_secondary"]).pack(anchor='w',pady=(0,4))
        tk.Label(body,text="Ingrese nueva fecha:",font=(FONT,10,'bold'),
                 bg=THEME["ct_bg"],fg=THEME["txt_secondary"]).pack(anchor='w',pady=(0,4))
        tk.Label(body,text="Formato: DD/MM/AAAA  (ej: 15/04/2026)",font=(FONT,8),
                 bg=THEME["ct_bg"],fg=THEME["txt_secondary"]).pack(anchor='w',pady=(0,8))

        date_outer=tk.Frame(body,bg=THEME["input_brd"])
        date_outer.pack(fill='x')
        date_inner=tk.Frame(date_outer,bg=THEME["input_bg"])
        date_inner.pack(fill='x',padx=1,pady=1)
        date_e=tk.Entry(date_inner,font=(FONT,14),bg=THEME["input_bg"],
                        fg=THEME["txt_primary"],relief='flat',bd=0,
                        justify='center',insertbackground=THEME["acc_blue"])
        date_e.pack(fill='x',padx=12,pady=10)
        date_outer.bind('<FocusIn>',  lambda e: date_outer.config(bg=THEME["acc_blue"]))
        date_e.bind('<FocusIn>',  lambda e: date_outer.config(bg=THEME["acc_blue"]))
        date_e.bind('<FocusOut>', lambda e: date_outer.config(bg=THEME["input_brd"]))
        date_e.insert(0, fecha_actual_display)
        date_e.select_range(0,'end')
        date_e.focus()

        lbl_err = tk.Label(body, text="", font=(FONT,9), bg=THEME["ct_bg"], fg=THEME["acc_rose"])
        lbl_err.pack(anchor='w', pady=(6,0))

        def update():
            raw = date_e.get().strip()
            try:
                from datetime import datetime as _dt
                # Aceptar DD/MM/AAAA o DD-MM-AAAA
                raw2 = raw.replace('-','/')
                parsed = _dt.strptime(raw2, '%d/%m/%Y')
                iso = parsed.strftime('%Y-%m-%d')
                self.db.update_next_payment_date(self.selected_credit_id, iso)
                messagebox.showinfo("✅","Fecha actualizada correctamente",parent=dlg)
                self.load_credits(); dlg.destroy()
            except ValueError:
                lbl_err.config(text="⚠  Formato inválido. Use DD/MM/AAAA  (ej: 15/04/2026)")
            except Exception as e:
                messagebox.showerror("Error",str(e),parent=dlg)

        date_e.bind('<Return>', lambda e: update())

        btn_act = _btn(btn_bar,"Actualizar",update,THEME["acc_green"],"💾")
        btn_act.enable()  # ← habilitar explícitamente
        btn_act.pack(side='left',padx=(0,8))
        btn_can = _btn(btn_bar,"Cancelar",dlg.destroy,THEME["btn_secondary"],"✕")
        btn_can.enable()  # ← habilitar explícitamente
        btn_can.pack(side='left')
